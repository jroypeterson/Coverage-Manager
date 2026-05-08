"""Excel report generation for performance reports."""

import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reporting.calcs import (
    RETURN_COLS, PERIOD_COLS, ANNUAL_COLS, FUND_COLS, VAL_COLS,
    FUND_PCT_COLS, FUND_MONEY_COLS, FUND_DISPLAY_NAMES,
    HIST_COLS, HIST_RATIO_COLS, HIST_VS_AVG_COLS,
    get_color, format_mkt_cap, format_price,
)


def write_excel_sheet(wb, sheet_name, df, info_cols):
    """Write a formatted Excel sheet for a given segment DataFrame."""
    ws = wb.create_sheet(title=sheet_name)

    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )
    fund_header_fill = PatternFill(start_color="1A5276", end_color="1A5276", fill_type="solid")
    val_header_fill = PatternFill(start_color="1A5276", end_color="1A5276", fill_type="solid")
    hist_header_fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
    etf_info_fill = PatternFill(start_color="D6DEE5", end_color="D6DEE5", fill_type="solid")
    etf_font_bold = Font(bold=True, size=9)
    etf_top_border = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="medium", color="2C3E50"),
        bottom=Side(style="thin", color="D0D0D0"),
    )

    all_cols = info_cols + RETURN_COLS + FUND_COLS + HIST_COLS

    # Write headers
    for col_idx, col_name in enumerate(all_cols, 1):
        display_name = FUND_DISPLAY_NAMES.get(col_name, col_name)
        cell = ws.cell(row=1, column=col_idx, value=display_name)
        if col_name in HIST_COLS:
            cell.fill = hist_header_fill
        elif col_name in FUND_COLS:
            cell.fill = fund_header_fill
        elif col_name in VAL_COLS:
            cell.fill = val_header_fill
        else:
            cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border

    # Write data
    val_and_fund_cols = set(FUND_COLS + VAL_COLS)
    hist_col_set = set(HIST_COLS)
    prev_was_etf = False
    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        is_etf = row.get("_is_etf", False)
        is_first_etf = is_etf and not prev_was_etf
        prev_was_etf = is_etf
        for col_idx, col_name in enumerate(all_cols, 1):
            val = row.get(col_name)
            if col_name in hist_col_set:
                if val is None or (hasattr(val, "__float__") and pd.isna(val)):
                    cell = ws.cell(row=row_idx, column=col_idx, value="N/A")
                    cell.font = Font(color="999999", size=9)
                elif col_name in HIST_VS_AVG_COLS:
                    try:
                        num_val = float(val)
                        cell = ws.cell(row=row_idx, column=col_idx, value=round(num_val, 1))
                        cell.number_format = '0.0"%"'
                        # Premium (positive) = red (expensive); discount (negative) = green (cheap).
                        # Invert get_color by flipping the sign.
                        hex_color = get_color(-num_val)
                        cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
                        if num_val > 0:
                            cell.font = Font(color="8B0000", size=9)
                        elif num_val < 0:
                            cell.font = Font(color="006400", size=9)
                        else:
                            cell.font = Font(size=9)
                    except (TypeError, ValueError):
                        cell = ws.cell(row=row_idx, column=col_idx, value=str(val))
                        cell.font = Font(size=9)
                else:
                    # Plain ratio (avg, +1σ, -1σ, min, max) — 1 decimal, no color
                    try:
                        cell = ws.cell(row=row_idx, column=col_idx, value=round(float(val), 1))
                        cell.number_format = '0.0'
                    except (TypeError, ValueError):
                        cell = ws.cell(row=row_idx, column=col_idx, value=str(val))
                    cell.font = Font(size=9)
            elif col_name in RETURN_COLS:
                if val is not None and not pd.isna(val):
                    cell = ws.cell(row=row_idx, column=col_idx, value=round(val, 1))
                    cell.number_format = '0.0"%"'
                    hex_color = get_color(val)
                    cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
                    if val < 0:
                        cell.font = Font(color="8B0000", size=9)
                    elif val > 0:
                        cell.font = Font(color="006400", size=9)
                    else:
                        cell.font = Font(size=9)
                else:
                    cell = ws.cell(row=row_idx, column=col_idx, value="N/A")
                    cell.font = Font(color="999999", size=9)
            elif col_name in val_and_fund_cols:
                needs_asterisk = False
                if col_name == "Rev Grw" and not row.get("_is_ttm_rev", False):
                    needs_asterisk = True
                elif col_name == "EPS Grw" and not row.get("_is_ttm_eps", False):
                    needs_asterisk = True
                if col_name in FUND_MONEY_COLS:
                    currency = row.get("_currency", "")
                    cell = ws.cell(row=row_idx, column=col_idx, value=format_mkt_cap(val, currency))
                    cell.font = Font(size=9)
                elif col_name == "Price":
                    currency = row.get("_currency", "")
                    cell = ws.cell(row=row_idx, column=col_idx, value=format_price(val, currency))
                    cell.font = Font(size=9)
                elif col_name == "% 52Wk Hi":
                    if val is not None and not pd.isna(val):
                        pct = float(val)
                        cell = ws.cell(row=row_idx, column=col_idx, value=round(pct, 1))
                        cell.number_format = '0.0"%"'
                        hex_color = get_color(pct - 100)
                        cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
                        cell.font = Font(size=9)
                    else:
                        cell = ws.cell(row=row_idx, column=col_idx, value="N/A")
                        cell.font = Font(color="999999", size=9)
                elif val is not None and not pd.isna(val):
                    if col_name in FUND_PCT_COLS:
                        asterisk = "*" if needs_asterisk else ""
                        try:
                            num_val = float(val)
                            cell = ws.cell(row=row_idx, column=col_idx, value=f"{num_val:.1f}%{asterisk}")
                            cell.number_format = '@'
                            hex_color = get_color(num_val)
                            cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
                            if num_val < 0:
                                cell.font = Font(color="8B0000", size=9)
                            elif num_val > 0:
                                cell.font = Font(color="006400", size=9)
                            else:
                                cell.font = Font(size=9)
                        except (TypeError, ValueError):
                            cell = ws.cell(row=row_idx, column=col_idx, value=str(val))
                            cell.font = Font(size=9)
                    else:
                        try:
                            cell = ws.cell(row=row_idx, column=col_idx, value=round(float(val), 1))
                            cell.number_format = '0.0'
                        except (TypeError, ValueError):
                            cell = ws.cell(row=row_idx, column=col_idx, value=str(val))
                        cell.font = Font(size=9)
                else:
                    cell = ws.cell(row=row_idx, column=col_idx, value="N/A")
                    cell.font = Font(color="999999", size=9)
            else:
                display_val = val if val and str(val) != "nan" else ""
                cell = ws.cell(row=row_idx, column=col_idx, value=display_val)
                cell.font = Font(size=9)

            # ETF benchmark row styling
            if is_etf and col_name not in RETURN_COLS:
                cell.fill = etf_info_fill
                cell.font = etf_font_bold
            if is_first_etf:
                cell.border = etf_top_border
            else:
                cell.border = thin_border
            cell.alignment = Alignment(horizontal="center" if col_name in RETURN_COLS or col_name in val_and_fund_cols else "left", vertical="center")

    # Column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 30
    val_widths = {"Mkt Cap": 13, "Enterprise Value": 15, "Net Debt": 13, "Price": 10, "% 52Wk Hi": 11}
    for i, vc in enumerate(VAL_COLS):
        ws.column_dimensions[get_column_letter(3 + i)].width = val_widths.get(vc, 12)
    meta_start = 3 + len(VAL_COLS)
    meta_widths = [18, 22, 20, 24, 8, 12]
    for i, w in enumerate(meta_widths):
        ws.column_dimensions[get_column_letter(meta_start + i)].width = w
    ret_start = meta_start + len(meta_widths)
    for ci in range(ret_start, ret_start + len(RETURN_COLS)):
        ws.column_dimensions[get_column_letter(ci)].width = 10
    fund_start = ret_start + len(RETURN_COLS)
    for i in range(len(FUND_COLS)):
        ws.column_dimensions[get_column_letter(fund_start + i)].width = 10
    hist_start = fund_start + len(FUND_COLS)
    for i in range(len(HIST_COLS)):
        ws.column_dimensions[get_column_letter(hist_start + i)].width = 11

    # Freeze panes
    freeze_col = get_column_letter(3 + len(VAL_COLS))
    ws.freeze_panes = f"{freeze_col}2"

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(all_cols))}{len(df) + 1}"

    # Group annual return columns
    annual_start_idx = len(info_cols) + len(PERIOD_COLS) + 1
    annual_end_idx = annual_start_idx + len(ANNUAL_COLS) - 1
    ws.column_dimensions.group(
        get_column_letter(annual_start_idx),
        get_column_letter(annual_end_idx),
        hidden=False,
    )

    # Footnote
    footnote_row = len(df) + 3
    cell = ws.cell(row=footnote_row, column=1,
                   value="* Value reflects quarterly YoY growth (yfinance) rather than TTM YoY (Finnhub). TTM YoY data was not available for this ticker.")
    cell.font = Font(size=8, italic=True, color="888888")
    ws.merge_cells(start_row=footnote_row, start_column=1, end_row=footnote_row, end_column=8)
