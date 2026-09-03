"""Build JP's Healthcare Services + MedTech coverage workbook.

One CURRENT file lives at `AA_Core Coverage.xlsx`; the previous
current file is moved into `archive/` stamped with the date it was built, so the
folder never accumulates dated copies at the top level. A stable filename is the
point -- it can be bookmarked, and the Google Sheet mirror keeps one identity.

Reads `exports/universe.csv` (the published artifact, never CM's internals) and
prices every row at Yahoo. Emits the workbook plus a flat CSV of the same rows
for the Google Sheet mirror.

    python scripts/build_hc_coverage_xlsx.py [--out-dir DIR] [--no-archive]

SYMBOLS GO THROUGH `ticker_utils.normalize_ticker`, NOT A PRIVATE ALIAS TABLE.
Coverage Manager's `Ticker` is correct by its own convention and is not rewritten
to suit a vendor -- `SHL GY` is how that row is keyed, and OpenFIGI confirms
ROG.SW is the Roche Genussschein. Yahoo needs a different string for many of
them, and `normalize_ticker` is where that translation already lives: the
space-suffix rule (`SHL GY` -> `SHL.DE`), `MANUAL_TICKER_MAP` for hyphenated
Nordic share classes (`COLOB DC` -> `COLO-B.CO`), and the Exchange fallback for
bare foreign symbols.

That last one is load-bearing rather than cosmetic. `MED` and `MOVE` are bare
SIX symbols that COLLIDE with live US listings, so a raw lookup returns Medifast
and Corvex -- a wrong company, not a missing one, which is the failure mode that
cost three repair passes in July and August 2026. `normalize_ticker` reads their
`Exchange` and yields `MED.SW` / `MOVE.SW`. Pass it the whole row, never just the
ticker string; it needs `Company Name` and `Exchange` to do either job.
"""
import argparse, csv, datetime, json, os, shutil, sys, tempfile, time
from concurrent.futures import ThreadPoolExecutor

import openpyxl
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
UNIVERSE = os.path.join(REPO, "exports", "universe.csv")
RATINGS_DIR = r"C:\Users\jroyp\Dropbox\Companies_Stocks_Sectors_Ratings"
# Moved here from Career\Pitches\Coverage on 2026-08-26 (JP), so the workbook,
# its archive and the ratings workbook it joins all live under one root.
DEFAULT_OUT = os.path.join(RATINGS_DIR, "_Coverage")
STEM = "AA_Core Coverage"


def dated_stem(day=None):
    """`AA_Core Coverage auto-updated - 08.26.26`. JP asked for the date in the
    filename, and "auto-updated" earns its place: this folder is his too, and the
    word is what tells him at a glance which files a machine owns."""
    d = day or datetime.date.today()
    return "%s auto-updated - %s" % (STEM, d.strftime("%m.%d.%y"))


# ⛑ ONLY THESE ARE OURS TO MOVE. JP 2026-08-26: "I might put my own files in this
# coverage folder for different reasons. Don't move my files. You just archive the
# files you auto-generate in the folder but leave the ones I put there manually
# alone." So archiving matches this allow-list of names THIS SCRIPT has ever
# produced -- never "every xlsx in the folder", which would sweep his work into
# archive/ the first time he dropped a file here.
AUTO_GENERATED_GLOBS = (
    "AA_Core Coverage auto-updated - *.xlsx",
    "AA_Core Coverage auto-updated - *.csv",
    "AA_Core Coverage.xlsx",                      # stable-named era, to 2026-08-26
    "AA_Core Coverage.csv",
    "Coverage - HC Services and MedTech.xlsx",    # pre-rename era
    "Coverage - HC Services and MedTech.csv",
)

# NOT renamed alongside the workbook. The Google Sheet mirror is a single
# =IMPORTDATA() cell pointed at this exact URL and nothing here can rewrite that
# cell, so renaming the endpoint silently empties the Sheet.
PUBLIC_CSV = os.path.join(REPO, "docs", "hc_coverage.csv")

RATINGS_PATH = os.path.join(RATINGS_DIR, "Ratings_CoreCoverage.xlsx")
# Scope: rows flagged `Core` in the universe -- the names JP covers analytically
# (310 of 1,346, spanning every sector, not just the HC segment). JP 2026-08-26:
# "I just want stocks in there that are coded as part of core=Y in the coverage
# manager." Seeding the whole universe made a 1,346-row sheet nobody would fill in.

SECTORS = ("Healthcare Services", "MedTech")

# ⛑ SCOPE IS NOT `Sector (JP)` ALONE, and the reason is a near-miss.
# On 2026-09-02 ARE, DOC, VTR and WELL moved to `Sector (JP)` = Real Estate so the
# universe agrees with GICS (Health Care REITs; ARE is Office REITs). That is a
# correct taxonomy fix and it would have silently dropped four names JP covers out
# of this workbook -- and, through `docs/hc_coverage.csv`, out of his Google Sheet
# -- on the next Friday build, with nothing anywhere reporting a change. JP:
# *"I don't want those names to drop out of coverage list AA_Coverage."*
# On 2026-09-03 the remaining 15 REITs in the subsector followed, so this clause
# now carries NINETEEN rows rather than four -- which is the clearest evidence it
# belongs here: had the fix been a per-name exception, the second migration would
# have re-broken the book the day after the first was patched.
# The distinction that resolves it: the GICS sector says which market an issuer
# TRADES in, the subsector says what it IS. This workbook is about the latter, so
# scope reads both. Any future sector re-map of a healthcare name must add its
# subsector here or it leaves the book unannounced.
SCOPE_SUBSECTORS = ("Healthcare Real Estate",)

# Human-readable form of the same rule, used in the provenance line and the
# Summary source note so the file states its own scope. One string, because two
# copies would drift apart the first time the rule changed.
SCOPE_DESCRIPTION = ("Sector (JP) in (Healthcare Services, MedTech), plus any row "
                     "whose Subsector (JP) is Healthcare Real Estate")


def in_scope(row):
    """True when a universe row belongs in this workbook.

    Takes a raw `exports/universe.csv` row (dict), not a built record.
    """
    return ((row.get("Sector (JP)") or "").strip() in SECTORS
            or (row.get("Subsector (JP)") or "").strip() in SCOPE_SUBSECTORS)


def split_sheets(recs):
    """Split built records into the (MedTech, everything-else) sheet buckets.

    `hs` is deliberately "not MedTech" rather than "== Healthcare Services": the
    two sheets and the Summary's two blocks must add up to the Coverage List, so a
    row admitted by SCOPE_SUBSECTORS under some third sector still has to land on
    one of them. Grouping is by subsector inside each block, so the healthcare
    REITs sit with the sixteen already there.
    """
    mt = [r for r in recs if r["Sector"] == "MedTech"]
    hs = [r for r in recs if r["Sector"] != "MedTech"]
    return mt, hs


# Calendar-year total returns, read from Coverage Manager's weekly performance
# snapshot rather than recomputed. JP, 2026-08-26: "You can just use a footnote to
# note when the returns are as of instead of re-running all the returns just for
# this report."
# Most recent on the LEFT, oldest on the right (JP). The snapshot names them
# plainly; `SNAPSHOT_RETURN_KEY` maps our heading back to its column.
CALENDAR_RETURNS = ["YTD", "2025", "2024", "2023", "2022", "2021", "2020", "2019"]
# ⛑ ANNUALISED, and the label says so. The snapshot's `3Y`/`5Y` are CUMULATIVE --
# `calc_period_return(hist, 365*3)` -- so JNJ's 3Y reads 74.4, which is 74% over
# three years and NOT 74% a year. JP asked for annual returns, so these are
# converted to a CAGR here. Publishing the raw figure under a heading saying
# "annual" would overstate a three-year return by roughly three times.
ANNUALISED_RETURNS = ["3Y ann.", "5Y ann."]
ANNUALISED_YEARS = {"3Y ann.": 3, "5Y ann.": 5}
ANNUALISED_SOURCE = {"3Y ann.": "3Y", "5Y ann.": "5Y"}
RETURN_COLS = CALENDAR_RETURNS + ANNUALISED_RETURNS
SNAPSHOT_MAX_AGE_DAYS = 10

# LC/SMID boundary. CONFIRMED BY JP 2026-08-26 ("Size threshold is fine"), which
# is what makes it a decision rather than an assumption -- it was NOT derived from
# his reference sheet, which only bounds the line to the open interval
# (USD 22.7bn SMID, USD 34.2bn LC] with no observation in between. It does
# reproduce that sheet on both sides: Guardant 22.6bn SMID, Illumina 34.1bn LC.
# Names near the line will flip with price and FX; the method is stated in the
# workbook so a flip reads as the rule working rather than as an error.
LC_THRESHOLD_USD_M = 25000

FIELDS = ["marketCap", "enterpriseValue", "regularMarketPrice", "currentPrice",
          "currency", "forwardPE", "fiftyTwoWeekHigh", "enterpriseToEbitda",
          "enterpriseToRevenue", "longName"]

# Order is JP's, 2026-08-26: size before sector, market cap and EV up front,
# venue information last, performance most-recent-first. `Rating` keeps the slot
# immediately after Company Name that he asked for earlier in the day -- he asked
# for market cap there too, so the two share the front and Rating leads.
COLS = (["Ticker", "Company Name", "Rating", "Mkt Cap (USD $M)", "EV (USD $M)",
         "Size", "Sector", "Subsector", "Sub-subsector", "Core Coverage",
         "Ccy", "Price (local)", "% of 52W High",
         "Fwd P/E (NTM)", "EV/Sales (TTM)", "EV/EBITDA (TTM)"]
        + RETURN_COLS
        + ["Listing", "Exchange", "Country (HQ)"])

# `docs/hc_coverage.csv` is served by GitHub Pages to ANYONE WITH THE URL, and it
# is what the Google Sheet reads. Anything in COLS is published unless it is named
# here. `Rating` was withheld until 2026-08-26, when JP asked for it in the Google
# file as well as the local one -- so his ratings are now publicly readable at that
# URL. That is his explicit call; keep the mechanism for the next sensitive column.
PRIVATE_ONLY = set()
PUBLIC_COLS = [c for c in COLS if c not in PRIVATE_ONLY]

WIDTH = {"Ticker": 11, "Company Name": 36, "Sector": 19, "Subsector": 26,
         "Sub-subsector": 20, "Core Coverage": 9, "Rating": 9, "Listing": 17,
         "Exchange": 16, "Country (HQ)": 15, "Ccy": 6, "Price (local)": 12,
         "Mkt Cap (USD $M)": 15, "EV (USD $M)": 14, "Size": 7,
         "% of 52W High": 12, "Fwd P/E (NTM)": 12,
         "EV/Sales (TTM)": 13, "EV/EBITDA (TTM)": 14}
WIDTH.update({c: 9 for c in RETURN_COLS})

HDR_FILL = PatternFill("solid", fgColor="1F3864")
RAMP_FILL = PatternFill("solid", fgColor="2E6B4F")
RAMP_CELL = PatternFill("solid", fgColor="E2EFDA")
HDR_FONT = Font(color="FFFFFF", bold=True, size=10)
TITLE_FONT = Font(bold=True, size=13, color="1F3864")
SUB_FONT = Font(size=9, italic=True, color="595959")
THIN = Side(style="thin", color="BFBFBF")
SUMHDR = ["Subsector", "Companies", "Core", "Rated", "Total Mkt Cap (USD $M)"]

# Decimal places per column, applied to BOTH the workbook's number format and the
# CSV's rounding, so the two surfaces cannot disagree about precision. JP
# 2026-08-26: "the annual returns dont need decimal point precision. and the
# Fwd P/e doesnt as well". Market cap has always been whole millions; the commas
# are display only, never in the CSV, where they would make the value text and
# break sorting in the Sheet.
DECIMALS = {"Mkt Cap (USD $M)": 0, "EV (USD $M)": 0, "Fwd P/E (NTM)": 0,
            "Price (local)": 2, "% of 52W High": 0,
            "EV/Sales (TTM)": 1, "EV/EBITDA (TTM)": 1}
DECIMALS.update({c: 0 for c in RETURN_COLS})
NUMFMT = {"Mkt Cap (USD $M)": "#,##0", "EV (USD $M)": "#,##0",
          "Fwd P/E (NTM)": "0", "Price (local)": "#,##0.00",
          "% of 52W High": '0"%"',
          "EV/Sales (TTM)": "0.0", "EV/EBITDA (TTM)": "0.0"}
NUMFMT.update({c: "0" for c in RETURN_COLS})


def num(x):
    try:
        f = float(x)
        return None if f != f else f
    except (TypeError, ValueError):
        return None


def size_bucket(mcap_usd_m):
    """LC / SMID / blank. Blank when the market cap is unknown -- the partial-book
    guard tolerates up to 5% missing, so a bucket must never be invented."""
    if mcap_usd_m is None:
        return None
    return "LC" if mcap_usd_m >= LC_THRESHOLD_USD_M else "SMID"


def pct_of_high(price, high):
    """Where the price sits in its own 52-week range, as a percent of the high.

    Blank unless both sides are real and the high is positive -- dividing by a
    missing or zero high yields either a crash or an infinity, and this column
    gets sorted."""
    p, h = num(price), num(high)
    if p is None or h is None or h <= 0:
        return None
    return 100.0 * p / h


def positive_multiple(raw):
    """A valuation multiple, or None if it is not positive.

    Same rule as `forward_pe` and for the same reason: a negative EV/EBITDA is a
    loss-making denominator showing through the ratio, not a cheap company, and it
    sorts straight to the top of any "cheapest names" ranking.
    """
    v = num(raw)
    return v if (v is not None and v > 0) else None


def forward_pe(raw):
    """Yahoo's forwardPE, or None. Non-positive is None, not a negative multiple.

    Yahoo returns a negative forwardPE for a company expected to lose money
    (Guardant: -380.9). That is the sign of the EPS estimate leaking through a
    ratio, not a valuation, and sorting a column containing it puts the biggest
    loss-maker at the top as if it were the cheapest name on the sheet.
    """
    v = num(raw)
    return v if (v is not None and v > 0) else None


def _payload_names_match(a, b):
    """Do two company names describe the same issuer?

    Reuses `universe.enrich._payload_names_match`, which is token-based precisely
    because character similarity cannot do this job -- Medartis/Medifast scores
    0.62 on difflib and they are different companies. Accents are stripped first:
    JP types these names by hand, so 'bioMerieux' must match 'bioMérieux'.
    Returns True when the comparison cannot be made (one side blank) -- an absent
    name is not evidence of a mismatch.
    """
    import unicodedata

    def strip_accents(s):
        return "".join(c for c in unicodedata.normalize("NFKD", str(s or ""))
                       if not unicodedata.combining(c))

    sys.path.insert(0, REPO)
    from universe.enrich import _payload_names_match as _match
    verdict = _match(strip_accents(a), strip_accents(b))
    return True if verdict is None else bool(verdict)


def fetch(rows):
    """rows: the universe dicts. Keyed by CM Ticker, valued by the Yahoo answer."""
    import yfinance as yf
    sys.path.insert(0, REPO)
    from ticker_utils import normalize_ticker

    symbols = {}
    for r in rows:
        sym = normalize_ticker(r["Ticker"], r.get("Company Name", ""), r.get("Exchange", ""))
        symbols[r["Ticker"]] = sym or r["Ticker"]

    def grab(cm):
        sym = symbols[cm]
        d = {}
        # Yahoo rate-limits aggressively once a session has pulled a few hundred
        # symbols, and a rate-limited response looks exactly like a company with
        # no market cap. Back off rather than accept the blank.
        for attempt in range(4):
            try:
                # ONE payload per attempt, then project. This was written as
                # `{k: yf.Ticker(sym).info.get(k) for k in FIELDS}`, which
                # re-fetches `.info` once PER FIELD -- 7 fields x 4 attempts x 239
                # tickers is up to 6,692 requests where 239 would do. That is what
                # throttled Yahoo hard enough to refuse 143 of 239 rows and send
                # this build to the FMP fallback three times in a row.
                info = yf.Ticker(sym).info or {}
                d = {k: info.get(k) for k in FIELDS}
                if d.get("marketCap"):
                    break
            except Exception as e:
                d = {"_error": str(e)[:90]}
            if attempt < 3:
                time.sleep(2 ** attempt)
        if not d.get("marketCap"):
            try:
                mc = getattr(yf.Ticker(sym).fast_info, "market_cap", None)
                if mc:
                    d["marketCap"] = float(mc)
            except Exception:
                pass
        d["_yf_symbol"] = sym
        return cm, d

    out = {}
    with ThreadPoolExecutor(max_workers=4) as ex:
        for cm, d in ex.map(grab, list(symbols)):
            out[cm] = d

    # FMP for whatever Yahoo would not answer. Yahoo throttles a session that has
    # pulled a few hundred symbols and then stays throttled for a long while, which
    # is a bad single point of failure for a file that is supposed to stay current;
    # FMP is a 300/min tier and answered every row Yahoo refused when this was added.
    #
    # It is given the SAME normalized symbol. Note what this deliberately does NOT
    # do: `provider_chain`'s AlphaVantage fallback strips the suffix
    # (`ticker.split(".")[0]`), which turns MED.SW back into MED and hands you
    # Medifast -- re-introducing the wrong-company bug on a correctly-keyed row.
    # That is why this calls the FMP provider directly rather than the chain.
    missing = [cm for cm, d in out.items() if not d.get("marketCap")]
    if not missing:
        return out
    try:
        from config import API_KEYS
        from providers.fmp_provider import fetch_fundamentals as fmp_fetch
        key = API_KEYS.get("FMP_API_KEY", "")
    except Exception as e:
        print("  FMP fallback unavailable: %s" % str(e)[:70], file=sys.stderr)
        return out
    if not key:
        print("  FMP fallback skipped: no FMP_API_KEY", file=sys.stderr)
        return out

    print("  Yahoo gave no market cap for %d rows; trying FMP" % len(missing))

    def fmp_grab(cm):
        try:
            res, _is_ttm, ccy = fmp_fetch(symbols[cm], key, use_cache=True)
            return cm, res, ccy
        except Exception:
            return cm, None, None

    filled = 0
    with ThreadPoolExecutor(max_workers=8) as ex:
        for cm, res, ccy in ex.map(fmp_grab, missing):
            if not res or res.get("Mkt Cap") is None:
                continue
            d = out[cm]
            d["marketCap"] = res["Mkt Cap"]
            if res.get("Price") is not None:
                d["regularMarketPrice"] = res["Price"]
            if ccy:
                d["currency"] = ccy
            d["_source"] = "fmp"
            filled += 1
    print("  FMP filled %d of %d" % (filled, len(missing)))
    return out


def fetch_fx(currencies):
    """Rates via Coverage Manager's own fx_provider, which caches for 12h.

    Rolling a private FX fetcher here cost a build: 239 ticker lookups exhausted
    the Yahoo budget and the FX calls that followed were all throttled, so the
    run aborted with every price already in hand. CM's provider caches, so a
    same-day rebuild pays nothing. CALL THIS BEFORE THE TICKER SWEEP.

    GBp is the one currency it cannot answer -- there is no `GBpUSD=X` -- and it
    needs two different rates anyway. Yahoo quotes LSE PRICES in pence but
    reports those companies' MARKET CAP in whole pounds, so pence is right for
    one column and 100x wrong for the other. Market cap uses fx['GBP']; price
    uses fx['GBp'], derived here.
    """
    sys.path.insert(0, REPO)
    from providers.fx_provider import fetch_fx_rates

    wanted = {c for c in currencies if c and c != "USD"}
    ask = {("GBP" if c == "GBp" else c) for c in wanted}
    fx = dict(fetch_fx_rates(sorted(ask)))
    if "GBp" in wanted and fx.get("GBP"):
        fx["GBp"] = fx["GBP"] / 100.0
    for c in wanted:
        if c not in fx:
            fx[c] = None
            print("  FX FAILED for %s" % c, file=sys.stderr)
    return fx


def annualise(cumulative_pct, years):
    """Cumulative % over `years` -> compound annual %, or None.

    The snapshot stores CUMULATIVE multi-year returns (`calc_period_return(hist,
    365*3)`), so JNJ's 3Y is 74.4 meaning 74% across three years. Reported under a
    heading that says "annual" that would overstate the result roughly threefold.

    Returns None below -100%: a security cannot lose more than everything, and the
    root of a negative growth factor is not a real number. Treating that as 0 or
    passing the raw value through would put an impossible figure in a column
    people rank on.
    """
    v = num(cumulative_pct)
    if v is None:
        return None
    growth = 1.0 + v / 100.0
    if growth <= 0:
        return None
    return (growth ** (1.0 / years) - 1.0) * 100.0


def load_returns():
    """Calendar-year returns from the newest weekly performance snapshot.

    Returns (by_ticker, as_of_date). These are TOTAL returns off split- and
    dividend-adjusted prices, which may not match how JP's reference sheet
    computes its columns -- said on the sheet rather than assumed away.

    The snapshot's date is its BUILD date, not each security's last trading day;
    the per-series observation date is discarded before the pickle is written.
    So the footnote says "snapshot built <date>", which is the only claim the
    data actually supports.
    """
    import glob
    import pandas as pd

    paths = sorted(glob.glob(os.path.join(REPO, "cache", "perf", "perf_df_*.pkl")),
                   key=os.path.getmtime)
    if not paths:
        print("  no performance snapshot found; return columns left blank", file=sys.stderr)
        return {}, None
    path = paths[-1]
    as_of = datetime.date.fromtimestamp(os.path.getmtime(path))
    age = (datetime.date.today() - as_of).days
    if age > SNAPSHOT_MAX_AGE_DAYS:
        # Blank beats stale-and-unlabelled: a YTD from a month ago sitting beside a
        # live price reads as one consistent moment and is not one.
        print("  performance snapshot is %d days old (>%d); return columns left blank"
              % (age, SNAPSHOT_MAX_AGE_DAYS), file=sys.stderr)
        return {}, None
    df = pd.read_pickle(path).set_index("Ticker")
    df = df[~df.index.duplicated(keep="first")]
    out = {}
    for t, row in df.iterrows():
        vals = {c: num(row.get(c)) for c in CALENDAR_RETURNS if c in df.columns}
        for col, years in ANNUALISED_YEARS.items():
            cum = num(row.get(ANNUALISED_SOURCE[col]))
            vals[col] = annualise(cum, years)
        out[str(t)] = vals
    return out, as_of


def load_ratings():
    """JP's ratings, keyed by ticker. The build NEVER writes this file."""
    if not os.path.exists(RATINGS_PATH):
        print("  no ratings file at %s; Rating left blank" % RATINGS_PATH, file=sys.stderr)
        return {}
    try:
        ws = openpyxl.load_workbook(RATINGS_PATH, data_only=True)["Ratings"]
    except Exception as e:
        print("  ratings file unreadable (%s); Rating left blank" % str(e)[:70], file=sys.stderr)
        return {}
    hdr = [(c.value if c.value is None else str(c.value).strip())
           for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(hdr) if h}
    if "Ticker" not in idx:
        print("  ratings file has no Ticker column; Rating left blank", file=sys.stderr)
        return {}
    out, dupes = {}, []
    for row in ws.iter_rows(min_row=2, values_only=True):
        t = row[idx["Ticker"]]
        if not t:
            continue
        t = str(t).strip()
        if t in out:
            dupes.append(t)
            continue
        out[t] = {
            "Rating": row[idx["Rating"]] if "Rating" in idx and idx["Rating"] < len(row) else None,
            "Company Name": (row[idx["Company Name"]]
                             if "Company Name" in idx and idx["Company Name"] < len(row) else None),
        }
    if dupes:
        # A duplicated key fans a left-join out into extra rows. Refuse rather than
        # silently pick one; JP has to resolve which rating is his.
        raise SystemExit("ABORT: duplicate Ticker in %s: %s. Nothing written."
                         % (RATINGS_PATH, ", ".join(sorted(set(dupes))[:10])))
    return out


def build_records(asof):
    rows = [r for r in csv.DictReader(open(UNIVERSE, encoding="utf-8"))
            if in_scope(r)]
    print("in scope: %d rows" % len(rows))
    # FX FIRST. The ticker sweep is what exhausts the Yahoo budget, and a run
    # that has every price but no rate to convert it with is a wasted run.
    fx = fetch_fx([r["Currency"] for r in rows] + ["GBP"])
    yf_data = fetch(rows)
    # A handful of rows quote in a currency the universe CSV does not claim
    # (SHMZF was a US OTC line recorded as JPY). Top those up rather than
    # dropping the row -- the vendor's currency is the one its numbers are in.
    extra = {(d.get("currency") or "").strip() for d in yf_data.values()}
    extra = {c for c in extra if c and c not in fx}
    if extra:
        print("  currencies Yahoo used that the universe did not declare: %s"
              % ", ".join(sorted(extra)))
        fx.update({k: v for k, v in fetch_fx(sorted(extra)).items() if k not in fx})
    returns, returns_asof = load_returns()
    ratings = load_ratings()

    # ⛑ Ticker is the join key on JP's instruction ("Ticker is an identity but it
    # can be fuzzy so you need to check and verify with me if its too ambiguous").
    # So: join on ticker, but where the ratings file's Company Name disagrees with
    # the universe's, DO NOT attach the rating -- collect it and make JP adjudicate.
    # A warning alone is not enough; the value has to be unreachable, or the
    # renderer publishes one issuer's rating against another's row.
    ambiguous = []

    recs = []
    for r in rows:
        t = r["Ticker"]
        d = yf_data.get(t, {})
        ccy = (d.get("currency") or r["Currency"] or "USD").strip()
        rate = fx.get("GBP") if ccy == "GBp" else fx.get(ccy)
        mc = num(d.get("marketCap"))
        ev = num(d.get("enterpriseValue"))
        px = num(d.get("regularMarketPrice")) or num(d.get("currentPrice"))
        mcap_usd_m = (mc * rate / 1e6) if (mc and rate) else None

        rating = None
        rr = ratings.get(t)
        if rr is not None:
            rated_name = str(rr.get("Company Name") or "").strip()
            if rated_name and not _payload_names_match(rated_name, r["Company Name"]):
                ambiguous.append((t, r["Company Name"], rated_name))
            else:
                rating = rr.get("Rating")

        rec = {
            "Ticker": t,
            "Company Name": r["Company Name"],
            "Sector": r["Sector (JP)"],
            "Subsector": r["Subsector (JP)"],
            "Sub-subsector": r["Sub-subsector (JP)"],
            "Core Coverage": "Y" if r.get("Core", "").strip().upper() == "Y" else "",
            "Rating": rating,
            "Listing": r["Listing Type"],
            "Exchange": r["Exchange"],
            "Country (HQ)": r["Country (HQ)"],
            "Ccy": ccy,
            "Price (local)": px,
            "Mkt Cap (USD $M)": mcap_usd_m,
            # Same payload and the same FX rate as market cap, deliberately: an EV
            # taken from a different source could be in a different currency and
            # the ratio between the two columns would be quietly meaningless.
            "EV (USD $M)": (ev * rate / 1e6) if (ev and rate) else None,
            # Blank, never a guessed bucket, when the market cap is unknown. The
            # partial-book guard tolerates up to 5% missing, so this does happen.
            "Size": size_bucket(mcap_usd_m),
            # Yahoo's forwardPE only -- genuinely NTM. Never backfilled from FMP's
            # trailing P/E, and never from Price/FY1-estimate either: an annual FY1
            # figure is not NTM, and its currency and share basis are unvalidated
            # against the price (pence-vs-pounds alone is a 100x trap). A blank
            # here means "not known", which is a true statement.
            "Fwd P/E (NTM)": forward_pe(d.get("forwardPE")),
            "% of 52W High": pct_of_high(px, num(d.get("fiftyTwoWeekHigh"))),
            # Non-positive blanked for the same reason as Fwd P/E: a negative
            # EV/EBITDA is a loss-making denominator showing through, not a cheap
            # company, and it sorts to the top of any "cheapest" ranking.
            "EV/Sales (TTM)": positive_multiple(d.get("enterpriseToRevenue")),
            "EV/EBITDA (TTM)": positive_multiple(d.get("enterpriseToEbitda")),
        }
        rec.update({c: None for c in RETURN_COLS})
        rec.update(returns.get(t, {}))
        recs.append(rec)
    recs.sort(key=lambda r: (0 if r["Sector"] == "MedTech" else 1,
                             -(r["Mkt Cap (USD $M)"] or 0)))
    # REFUSE A PARTIAL BOOK. A rate-limited Yahoo response is indistinguishable
    # from a company that has no market cap, so a build that quietly drops a
    # third of the rows publishes a coverage list whose totals are simply wrong
    # -- and it overwrites the good file that was there. The first run of this
    # script hit exactly that: 105 of 239 rows blank, USD 3,089bn against a true
    # ~4,950bn, written without complaint beyond a warning line.
    dead_fx = [c for c, v in fx.items() if v is None]
    missing = [r["Ticker"] for r in recs if not r["Mkt Cap (USD $M)"]]
    if dead_fx:
        raise SystemExit("ABORT: no FX rate for %s. Nothing written; re-run when "
                         "the rate limit clears." % ", ".join(dead_fx))
    if len(missing) > max(5, 0.05 * len(recs)):
        raise SystemExit(
            "ABORT: %d of %d rows have no market cap (%s...). That is almost "
            "certainly rate limiting, not %d delisted companies. Nothing "
            "written; re-run when the rate limit clears."
            % (len(missing), len(recs), ", ".join(missing[:8]), len(missing)))
    if missing:
        print("  %d rows have no market cap: %s"
              % (len(missing), ", ".join(missing)), file=sys.stderr)

    # Cardinality gates. A join that matched NOTHING publishes an entirely blank
    # column and looks exactly like a column of honest blanks, which is how a
    # silently-broken join survives. Say it out loud instead.
    joined_returns = sum(1 for r in recs if any(r.get(c) is not None for c in RETURN_COLS))
    joined_ratings = sum(1 for r in recs if r.get("Rating") not in (None, ""))
    if returns and joined_returns == 0:
        raise SystemExit("ABORT: the performance snapshot matched 0 of %d rows. "
                         "Nothing written." % len(recs))
    print("  returns joined for %d/%d rows | ratings joined for %d"
          % (joined_returns, len(recs), joined_ratings))
    if ambiguous:
        print("  %d ticker(s) NOT rated - the ratings file names a different company:"
              % len(ambiguous), file=sys.stderr)
        for t, uni, rated in ambiguous:
            print("     %-10s universe=%-32s ratings=%s" % (t, uni[:32], rated),
                  file=sys.stderr)
    return recs, returns_asof, ambiguous


def write_sheet(wb, title, rows, subtitle):
    ws = wb.create_sheet(title)
    ws["A1"] = ("Healthcare Services & MedTech - Coverage List"
                if title == "Coverage List" else title)
    ws["A1"].font = TITLE_FONT
    ws["A2"] = subtitle
    ws["A2"].font = SUB_FONT
    hr = 4
    c = ws.cell(hr, 1, "#")
    c.font, c.fill = HDR_FONT, HDR_FILL
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for j, col in enumerate(COLS, start=2):
        c = ws.cell(hr, j, col)
        c.font = HDR_FONT
        c.fill = RAMP_FILL if col == "Rating" else HDR_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[hr].height = 30

    for i, r in enumerate(rows, start=1):
        rr = hr + i
        ws.cell(rr, 1, i).alignment = Alignment(horizontal="center")
        for j, col in enumerate(COLS, start=2):
            cell = ws.cell(rr, j, r[col])
            if col in NUMFMT:
                cell.number_format = NUMFMT[col]
            if col in ("Core Coverage", "Rating", "Ccy", "Size"):
                cell.alignment = Alignment(horizontal="center")
            if col == "Company Name":
                cell.font = Font(bold=True, size=10)
            if col == "Rating" and r[col] not in (None, ""):
                cell.fill = RAMP_CELL
                cell.font = Font(bold=True, size=10, color="1F5132")
        for j in range(1, len(COLS) + 2):
            ws.cell(rr, j).border = Border(bottom=THIN)

    # Red / white / green across each return column INDEPENDENTLY, with white
    # pinned to zero rather than to the column's median. A midpoint of "50th
    # percentile" would paint a column where everything fell as if half of it were
    # fine; anchoring at 0 means the colour always answers "did this make money",
    # and only the intensity is relative to the column. `min`/`max` keep the
    # gradient scaled per column, which is what makes 2022 readable next to 2021.
    if rows:
        first, last = hr + 1, hr + len(rows)
        for j, col in enumerate(COLS, start=2):
            if col not in RETURN_COLS:
                continue
            letter = get_column_letter(j)
            ws.conditional_formatting.add(
                "%s%d:%s%d" % (letter, first, letter, last),
                ColorScaleRule(
                    start_type="min", start_color="F8696B",     # red
                    mid_type="num", mid_value=0, mid_color="FFFFFF",
                    end_type="max", end_color="63BE7B"))        # green

    ws.freeze_panes = ws.cell(hr + 1, 4)
    ws.auto_filter.ref = "A%d:%s%d" % (hr, get_column_letter(len(COLS) + 1), hr + len(rows))
    ws.column_dimensions["A"].width = 5
    for j, col in enumerate(COLS, start=2):
        ws.column_dimensions[get_column_letter(j)].width = WIDTH[col]


def write_summary(wb, allr, mt, hs, asof, src):
    ws = wb.create_sheet("Summary")
    ws["A1"] = "Coverage Summary - Healthcare Services & MedTech"
    ws["A1"].font = TITLE_FONT
    nrated = sum(1 for r in allr if r["Rating"] not in (None, ""))
    ws["A2"] = ("As of %s.  %d companies (%d MedTech, %d Healthcare Services).  "
                "%d marked Core Coverage.  %d carry a Rating, %d blank."
                % (asof, len(allr), len(mt), len(hs),
                   sum(1 for r in allr if r["Core Coverage"] == "Y"), nrated,
                   len(allr) - nrated))
    ws["A2"].font = SUB_FONT

    row = 4
    for sec, rows in (("MedTech", mt), ("Healthcare Services", hs)):
        ws.cell(row, 1, sec).font = Font(bold=True, size=11, color="1F3864")
        row += 1
        for j, h in enumerate(SUMHDR, start=1):
            c = ws.cell(row, j, h)
            c.font = HDR_FONT
            c.fill = RAMP_FILL if h == "Rated" else HDR_FILL
            c.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.row_dimensions[row].height = 28
        row += 1
        subs = {}
        for r in rows:
            subs.setdefault(r["Subsector"] or "(unclassified)", []).append(r)
        for name, grp in sorted(subs.items(),
                                key=lambda kv: -sum(x["Mkt Cap (USD $M)"] or 0 for x in kv[1])):
            caps = [x["Mkt Cap (USD $M)"] for x in grp if x["Mkt Cap (USD $M)"]]
            ws.cell(row, 1, name)
            ws.cell(row, 2, len(grp)).alignment = Alignment(horizontal="center")
            ws.cell(row, 3, sum(1 for x in grp if x["Core Coverage"] == "Y")).alignment = Alignment(horizontal="center")
            c = ws.cell(row, 4, sum(1 for x in grp if x["Rating"] not in (None, "")))
            c.alignment = Alignment(horizontal="center")
            c.fill = RAMP_CELL
            ws.cell(row, 5, sum(caps) if caps else None).number_format = "#,##0"
            row += 1
        tot = [x["Mkt Cap (USD $M)"] for x in rows if x["Mkt Cap (USD $M)"]]
        ws.cell(row, 1, sec + " total").font = Font(bold=True)
        for j, v in ((2, len(rows)),
                     (3, sum(1 for x in rows if x["Core Coverage"] == "Y")),
                     (4, sum(1 for x in rows if x["Rating"] not in (None, "")))):
            c = ws.cell(row, j, v)
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center")
        c = ws.cell(row, 5, sum(tot))
        c.number_format, c.font = "#,##0", Font(bold=True)
        row += 3

    ws.cell(row, 1, "Notes").font = Font(bold=True, size=11, color="1F3864")
    for n in [
        src,
        "Rebuilt by Coverage Manager scripts/build_hc_coverage_xlsx.py. The previous version of this file is in archive/, stamped with the date it was built.",
        "Rating is joined by ticker from Ratings_CoreCoverage.xlsx in the parent folder, which this build never writes - it is yours to edit. A blank means you have not rated that name yet.",
        "A rating is NOT attached when that file names a different company for the same ticker; the row is reported instead, because a ticker can be reassigned to another issuer and a silently carried-over rating is worse than a blank one.",
        "Core Coverage = the 'Core' flag on the Coverage Manager universe - the names you cover analytically. Separate from whether you have rated them.",
        "Prices are in local trading currency; market cap is USD at spot FX on the build date. The LSE names quote in pence (GBp).",
        "FOOTNOTE ON THE RETURN COLUMNS (2019-2025, YTD): these are TOTAL returns on split- and dividend-adjusted prices, and they are NOT as of the build date above - they come from the Coverage Manager weekly performance snapshot named in the caption. YTD in particular is as of that snapshot, while Price and Mkt Cap are same-day. A blank is a company that did not trade that year, never a zero.",
        "Size: LC at or above USD 25,000M market cap, SMID below. That threshold is a policy choice, not a measurement - the reference sheet only pins it between USD 22.7bn (SMID) and USD 34.2bn (LC). Names near the line move with price and FX.",
        "EVERY VALUATION COLUMN STATES ITS BASIS IN THE HEADING. Fwd P/E (NTM) is Yahoo's forwardPE, next twelve months. EV/Sales (TTM) and EV/EBITDA (TTM) are trailing twelve months, from yfinance enterpriseToRevenue / enterpriseToEbitda, or FMP evToSalesTTM / enterpriseValueMultipleTTM where Yahoo had no answer - both trailing, verified 2026-08-26.",
        "Fwd P/E is never backfilled from a trailing P/E or an annual FY1 estimate; either would be a different measure under the same heading, which is the mistake Coverage Manager's own Fwd P/E column made. All three multiples are blank where non-positive: a negative EV/EBITDA is a loss-making denominator showing through, not a cheap company, and it sorts to the top of any cheapest-first ranking.",
        "% of 52W High is the current price over the 52-week high, so 100% means the name is at its high. It is NOT colour-scaled - the return columns centre on zero and these do not, so one scale cannot serve both.",
        "Rating is joined from Companies_Stocks_Sectors_Ratings/Ratings_CoreCoverage.xlsx (Core=Y names only), which this build never writes. It is omitted from the published CSV and the Google Sheet.",
        "Some rows are priced under a different symbol than the Ticker column shows - Coverage Manager keys them by its own convention and Yahoo needs another string. MED and MOVE are the ones that matter: their bare symbols collide with live US listings.",
    ]:
        row += 1
        c = ws.cell(row, 1, "- " + n)
        c.font = Font(size=9, color="404040")
        c.alignment = Alignment(vertical="top")
    for col, w in zip("ABCDE", (40, 12, 9, 10, 22)):
        ws.column_dimensions[col].width = w


def archive_previous_autogenerated(out_dir, keep_names):
    """Move OUR earlier outputs into archive/. Leave everything else alone.

    ⛑ JP keeps his own files in this folder. "Don't move my files. You just
    archive the files you auto-generate in the folder but leave the ones I put
    there manually alone." So this matches an explicit allow-list of names this
    script has produced (`AUTO_GENERATED_GLOBS`) rather than sweeping the folder
    by extension -- the lazy version would file his work under archive/ the first
    time he dropped a workbook here, and he would have no way to tell that from a
    file he had misplaced himself.

    `keep_names` is today's output, which must survive: with the date in the
    filename, a same-day rebuild would otherwise archive the file it just wrote.
    """
    import glob as _glob

    arch_dir = os.path.join(out_dir, "archive")
    moved = []
    keep = {n.lower() for n in keep_names}
    for pattern in AUTO_GENERATED_GLOBS:
        for path in _glob.glob(os.path.join(out_dir, pattern)):
            name = os.path.basename(path)
            if name.lower() in keep:
                continue
            os.makedirs(arch_dir, exist_ok=True)
            dest = os.path.join(arch_dir, name)
            n = 2
            root, ext = os.path.splitext(name)
            while os.path.exists(dest):
                dest = os.path.join(arch_dir, "%s (%d)%s" % (root, n, ext))
                n += 1
            shutil.move(path, dest)
            moved.append(os.path.basename(dest))
    return moved


RATING_SHEET = "Ratings"
RATING_COLS = ["Ticker", "Company Name", "Sector (JP)", "Subsector (JP)",
               "Rating", "Notes", "Status", "Last Synced"]
HUMAN_COLS = {"Rating", "Notes"}          # never written by any code path
MACHINE_COLS = {"Sector (JP)", "Subsector (JP)", "Status", "Last Synced"}


def sync_ratings():
    """Seed / refresh the ratings workbook. Only reachable via --sync-ratings.

    The ordinary build NEVER calls this: the file is JP's to type in, and a daily
    automated job writing the file he is editing is how edits get lost.

    Ownership is split and enforced:
      * `Rating` and `Notes` are HUMAN. No code path writes them, including this one.
      * `Sector (JP)`, `Subsector (JP)`, `Status`, `Last Synced` are MACHINE and are
        refreshed, because otherwise they rot as the universe moves.
      * `Ticker` and `Company Name` are the identity, and are the interesting case.

    ⛑ A CHANGED `Company Name` IS NOT REFRESHED. That looks like a bug and is the
    whole safety mechanism. The build attaches a rating by ticker only when the
    ratings file's company name still agrees with the universe's; if this function
    quietly rewrote the name whenever it drifted, the two would agree by
    construction and the check could never fire again -- which is precisely how
    `ZEN` kept Zendesk's classification after the ticker became Zentek. Instead the
    row is marked `REVIEW - issuer may have changed`, the rating stops attaching,
    and JP adjudicates. Fail closed.
    """
    everything = list(csv.DictReader(open(UNIVERSE, encoding="utf-8")))
    uni = [r for r in everything if r.get("Core", "").strip().upper() == "Y"]
    # Keyed on the FULL universe, not just the core slice: a row that has dropped
    # out of Core still needs its name checked before its rating is trusted, and
    # "no longer core" is a different fact from "no longer exists".
    by_ticker = {r["Ticker"]: r for r in everything}
    core_tickers = {r["Ticker"] for r in uni}
    today = datetime.date.today().isoformat()
    print("core=Y in universe: %d of %d rows" % (len(uni), len(everything)))

    if os.path.exists(RATINGS_PATH):
        wb = openpyxl.load_workbook(RATINGS_PATH)
        if RATING_SHEET not in wb.sheetnames:
            raise SystemExit("ABORT: %s has no '%s' sheet." % (RATINGS_PATH, RATING_SHEET))
        ws = wb[RATING_SHEET]
        hdr = [(c.value if c.value is None else str(c.value).strip())
               for c in next(ws.iter_rows(min_row=1, max_row=1))]
        idx = {h: i + 1 for i, h in enumerate(hdr) if h}
        for need in ("Ticker", "Rating"):
            if need not in idx:
                raise SystemExit("ABORT: %s is missing the '%s' column." % (RATINGS_PATH, need))
    else:
        os.makedirs(os.path.dirname(RATINGS_PATH), exist_ok=True)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = RATING_SHEET
        for j, c in enumerate(RATING_COLS, start=1):
            cell = ws.cell(1, j, c)
            cell.font = HDR_FONT
            cell.fill = RAMP_FILL if c in HUMAN_COLS else HDR_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 28
        idx = {c: j for j, c in enumerate(RATING_COLS, start=1)}

    seen, dupes, added, reviewed, retired, uncored = {}, [], 0, [], 0, 0
    for rr in range(2, ws.max_row + 1):
        t = ws.cell(rr, idx["Ticker"]).value
        if not t:
            continue
        t = str(t).strip()
        if t in seen:
            dupes.append(t)
            continue
        seen[t] = rr
    if dupes:
        raise SystemExit("ABORT: duplicate Ticker rows in %s: %s. Resolve by hand; "
                         "nothing written." % (RATINGS_PATH, ", ".join(sorted(set(dupes))[:10])))

    for t, rr in seen.items():
        u = by_ticker.get(t)
        if u is None:
            ws.cell(rr, idx["Status"], "Not in universe")
            ws.cell(rr, idx["Last Synced"], today)
            retired += 1
            continue
        if t not in core_tickers:
            # Dropped out of core coverage. The row and JP's rating STAY -- he formed
            # that view and it is not the sync's to discard -- but the status says so,
            # so a stale rating is visible rather than silently current.
            ws.cell(rr, idx["Status"], "No longer Core=Y")
            ws.cell(rr, idx["Last Synced"], today)
            uncored += 1
            continue
        stored_name = str(ws.cell(rr, idx["Company Name"]).value or "").strip()
        if stored_name and not _payload_names_match(stored_name, u["Company Name"]):
            # Leave EVERY cell alone except the flag. See the docstring.
            ws.cell(rr, idx["Status"], "REVIEW - issuer may have changed")
            reviewed.append((t, stored_name, u["Company Name"]))
            continue
        if not stored_name:
            ws.cell(rr, idx["Company Name"], u["Company Name"])
        ws.cell(rr, idx["Sector (JP)"], u["Sector (JP)"])
        ws.cell(rr, idx["Subsector (JP)"], u["Subsector (JP)"])
        ws.cell(rr, idx["Status"], "Active")
        ws.cell(rr, idx["Last Synced"], today)

    row = ws.max_row + 1
    for r in uni:
        if r["Ticker"] in seen:
            continue
        ws.cell(row, idx["Ticker"], r["Ticker"])
        ws.cell(row, idx["Company Name"], r["Company Name"])
        ws.cell(row, idx["Sector (JP)"], r["Sector (JP)"])
        ws.cell(row, idx["Subsector (JP)"], r["Subsector (JP)"])
        ws.cell(row, idx["Status"], "Active")
        ws.cell(row, idx["Last Synced"], today)
        row += 1
        added += 1

    ws.freeze_panes = "C2"
    ws.auto_filter.ref = "A1:%s%d" % (get_column_letter(len(RATING_COLS)), ws.max_row - 1)
    for c, w in zip("ABCDEFGH", (12, 38, 22, 26, 10, 46, 30, 13)):
        ws.column_dimensions[c].width = w

    try:
        wb.save(RATINGS_PATH)
    except PermissionError:
        raise SystemExit("ABORT: %s is open in Excel. Close it and re-run; nothing written."
                         % RATINGS_PATH)

    print("ratings file: %s" % RATINGS_PATH)
    print("  rows now %d | added %d | not-in-universe %d | no-longer-core %d | "
          "flagged for review %d"
          % (ws.max_row - 1, added, retired, uncored, len(reviewed)))
    for t, old, new in reviewed:
        print("     REVIEW %-10s ratings says '%s', universe says '%s'" % (t, old, new))
    if reviewed:
        print("  Those rows keep their rating in this file but the coverage build will"
              " NOT attach it until the name is reconciled.")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--out-dir", default=DEFAULT_OUT)
    ap.add_argument("--no-archive", action="store_true")
    ap.add_argument("--sync-ratings", action="store_true",
                    help="Seed/refresh the ratings workbook, then exit. Never run by "
                         "the ordinary build.")
    args = ap.parse_args()

    if args.sync_ratings:
        sync_ratings()
        return

    asof = datetime.date.today().isoformat()
    recs, returns_asof, ambiguous = build_records(asof)
    mt, hs = split_sheets(recs)
    if returns_asof:
        ret_note = ("Returns are from the Coverage Manager performance snapshot "
                    "built %s" % returns_asof.isoformat())
    else:
        ret_note = "Returns unavailable (no fresh performance snapshot)"
    src = ("Source: Coverage Manager exports/universe.csv, filtered to %s. Price and "
           "market cap pulled %s from Yahoo Finance, falling back to FMP per row "
           "where Yahoo had no answer. %s."
           % (SCOPE_DESCRIPTION, asof, ret_note))

    # One line that travels WITH the data. The xlsx carries it as the subtitle
    # under the title; the CSV carries it as a preamble row, which is the only way
    # it reaches the Google Sheet -- that Sheet is a single =IMPORTDATA cell and
    # nothing here can write to any other cell of it. JP asked for "when it was
    # last updated and any relevant background" to live in the file itself, so a
    # reader never has to come and ask which of these numbers is stale.
    provenance = (
        "AA_Core Coverage - LAST UPDATED %s %s. Rebuilt automatically every Friday "
        "by Coverage Manager (scheduled task WeeklyCoverageBuilder), immediately "
        "after the weekly performance run so prices and returns share one as-of "
        "date. %d names: %s. "
        "Price/market cap from Yahoo Finance that morning, FMP per row where Yahoo "
        "had no answer; market cap USD at spot FX, price in local currency. %s. "
        "Size: LC at or above USD 25,000M, else SMID. Valuation columns state their "
        "basis in the heading: Fwd P/E (NTM) forward, EV/Sales and EV/EBITDA trailing "
        "twelve months. Rating is joined by ticker "
        "from Ratings_CoreCoverage.xlsx and is yours to edit - this build never "
        "writes it."
        % (asof, datetime.datetime.now().strftime("%H:%M"), len(recs),
           SCOPE_DESCRIPTION, ret_note))

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    write_sheet(wb, "Coverage List", recs, provenance)
    write_sheet(wb, "MedTech", mt, "%d names.  %s" % (len(mt), provenance))
    write_sheet(wb, "Healthcare Services", hs, "%d names.  %s" % (len(hs), provenance))
    write_summary(wb, recs, mt, hs, asof, src)
    wb.move_sheet("Summary", offset=-4)

    # Save to a scratch path FIRST, so a failed write cannot leave the folder
    # with the old file already archived and no current file in its place.
    # A plain copy rather than os.replace: Dropbox holds a lock that makes
    # atomic replace fail with WinError 5.
    stem_today = dated_stem()
    current = os.path.join(args.out_dir, "%s.xlsx" % stem_today)
    private_csv = os.path.join(args.out_dir, "%s.csv" % stem_today)
    tmp = os.path.join(tempfile.gettempdir(), "hc_coverage_%d.xlsx" % os.getpid())
    wb.save(tmp)
    # The archive MOVE and the install COPY share one handler on purpose. Excel
    # holds a sharing lock on an open workbook, and the first thing that touches
    # the file is the archive step's `shutil.move` -- so the PermissionError
    # surfaces THERE, not at the copy below. Guarding only the copy left the move
    # to raise an unhandled traceback and exit 1, which is the red-task false
    # alarm the exit-3 code exists to avoid. (A read-only ATTRIBUTE does not
    # reproduce this: move succeeds on those, which is why the first attempt to
    # test this passed and proved nothing.)
    try:
        if not args.no_archive:
            # Today's two outputs are excluded, or a same-day rebuild would file
            # away the very files it is about to write.
            moved = archive_previous_autogenerated(
                args.out_dir,
                keep_names=[os.path.basename(current), os.path.basename(private_csv)])
            for name in moved:
                print("archived -> archive/%s" % name)
        shutil.copy2(tmp, current)
    except PermissionError:
        # EXIT 3 = "the file is open in Excel", deliberately distinct from every
        # other failure. The weekly scheduled build calls this, and JP having the
        # workbook open on a Friday morning is not a broken pipeline -- turning
        # the task RED for it would train him to ignore the red. Exit 1 stays for
        # the failures that ARE the pipeline's fault (dead FX, a partial book, an
        # unreadable ratings file), which no amount of re-running fixes on its own.
        print("ERROR: %s is open in Excel, so the workbook was NOT replaced. "
              "The new one is at %s and the previous is in archive/. Close Excel "
              "and re-run." % (current, tmp), file=sys.stderr)
        raise SystemExit(3)
    finally:
        if os.path.exists(current):
            os.remove(tmp)
    print("wrote %s" % current)

    # Flat CSV, written twice on purpose:
    #   1. beside the workbook, so the folder is self-contained
    #   2. into docs/, which GitHub Pages serves at a STABLE public URL. The
    #      Google Sheet mirror is one =IMPORTDATA() cell pointed at that URL, so
    #      it refreshes itself and keeps ONE file id forever. The Drive connector
    #      can only rename and move a Sheet -- it cannot write cells -- so the
    #      alternative was replacing the file every build and minting a new URL
    #      each time, which breaks every link to it.
    def _flatten(cols, preamble=None):
        out = []
        if preamble:
            out.append([preamble])
            out.append([])
        out.append(["#"] + cols)
        for i, r in enumerate(recs, 1):
            row = [i]
            for c in cols:
                v = r.get(c)
                if v is None:
                    row.append("")
                elif isinstance(v, float):
                    dp = DECIMALS.get(c, 2)
                    row.append(int(round(v)) if dp == 0 else round(v, dp))
                else:
                    row.append(v)
            out.append(row)
        return out

    # Beside the workbook: the FULL schema, ratings included. This folder is JP's.
    with open(private_csv, "w", newline="", encoding="utf-8") as fh:
        csv.writer(fh).writerows(_flatten(COLS, provenance))
    print("wrote %s" % private_csv)

    # docs/: PUBLIC. Served by GitHub Pages to anyone with the URL, and read by the
    # Google Sheet. Ratings are dropped here -- see PRIVATE_ONLY.
    os.makedirs(os.path.dirname(PUBLIC_CSV), exist_ok=True)
    with open(PUBLIC_CSV, "w", newline="", encoding="utf-8") as fh:
        csv.writer(fh).writerows(_flatten(PUBLIC_COLS, provenance))
    print("wrote %s  (public schema, %d of %d columns)"
          % (PUBLIC_CSV, len(PUBLIC_COLS), len(COLS)))
    print("  NOTE: docs/ is only served after a git commit+push of this repo.")

    print(json.dumps({
        "asof": asof, "rows": len(recs), "medtech": len(mt), "hc_services": len(hs),
        "rated": sum(1 for r in recs if r["Rating"] not in (None, "")),
        "total_mkt_cap_usd_bn": round(sum(r["Mkt Cap (USD $M)"] or 0 for r in recs) / 1000, 1),
    }, indent=1))


if __name__ == "__main__":
    main()
