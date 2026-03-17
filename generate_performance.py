import pandas as pd
import yfinance as yf
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, date
import warnings
import sys
import glob
import shutil
import os
import requests
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

from ticker_utils import normalize_ticker, MANUAL_TICKER_MAP, CSV_PATH, REPORTS_DIR, SCRIPT_DIR

warnings.filterwarnings("ignore")

# Load API keys from .env
ENV_PATH = SCRIPT_DIR / ".env"
API_KEYS = {}
if ENV_PATH.exists():
    with open(ENV_PATH) as f:
        for line in f:
            line = line.strip()
            if "=" in line and not line.startswith("#"):
                k, v = line.split("=", 1)
                API_KEYS[k.strip()] = v.strip()

TODAY = date.today().strftime("%Y-%m-%d")
OLD_REPORTS_DIR = REPORTS_DIR / "old reports"
OUTPUT_XLSX = REPORTS_DIR / f"coverage_performance_{TODAY}.xlsx"
OUTPUT_HTML = REPORTS_DIR / f"coverage_performance_{TODAY}.html"

ANNUAL_YEARS = list(range(2019, date.today().year + 1))
PERIOD_COLS = ["1D", "1W", "QTD", "YTD", "1Y", "3Y", "5Y", "10Y"]
RETURN_COLS = PERIOD_COLS + [str(y) for y in ANNUAL_YEARS]
FUND_COLS = ["Mkt Cap", "Fwd P/E", "EV/EBITDA", "PEG", "Gross Mgn", "Op Mgn", "ROE", "Rev Grw", "EPS Grw"]

# Formatting rules for fundamental columns
FUND_PCT_COLS = {"Gross Mgn", "Op Mgn", "ROE", "Rev Grw", "EPS Grw"}  # displayed as %
FUND_RATIO_COLS = {"Fwd P/E", "EV/EBITDA", "PEG"}  # displayed as ratio (1 decimal)
FUND_MONEY_COLS = {"Mkt Cap"}  # displayed as abbreviated dollar amount

# Display names for fundamental column headers (internal name -> display name)
FUND_DISPLAY_NAMES = {
    "Mkt Cap": "Mkt Cap",
    "Fwd P/E": "Fwd P/E",
    "EV/EBITDA": "EV/EBITDA",
    "PEG": "PEG",
    "Gross Mgn": "Gross Mgn (TTM)",
    "Op Mgn": "Op Mgn (TTM)",
    "ROE": "ROE (TTM)",
    "Rev Grw": "Rev Grw (TTM YoY)",
    "EPS Grw": "EPS Grw (TTM YoY)",
}


def try_fmp_historical(ticker, api_key):
    """Try to get historical prices from FMP API as fallback (US tickers only)."""
    try:
        url = f"https://financialmodelingprep.com/stable/historical-price-eod/full?symbol={ticker}&apikey={api_key}"
        resp = requests.get(url, timeout=10)
        if resp.status_code != 200:
            return None
        data = resp.json()
        if not data or not isinstance(data, list):
            return None
        df = pd.DataFrame(data)
        if "date" not in df.columns or "close" not in df.columns:
            return None
        df["date"] = pd.to_datetime(df["date"])
        df = df.set_index("date").sort_index()
        series = df["close"].dropna()
        return series if len(series) > 0 else None
    except Exception:
        return None


def fetch_finnhub_metrics(ticker, api_key):
    """Fetch fundamental metrics from Finnhub (US tickers only on free tier)."""
    try:
        url = f"https://finnhub.io/api/v1/stock/metric?symbol={ticker}&metric=all&token={api_key}"
        resp = requests.get(url, timeout=10)
        if resp.status_code != 200:
            return {}
        data = resp.json()
        return data.get("metric", {})
    except Exception:
        return {}


def fetch_fundamentals(yf_ticker, finnhub_metrics=None):
    """Fetch fundamental data from yfinance, enriched with Finnhub for US tickers.

    Returns (result, is_ttm) where is_ttm tracks whether Finnhub TTM YoY data
    was used for Rev Grw and EPS Grw.
    """
    result = {col: None for col in FUND_COLS}
    is_ttm = {"Rev Grw": False, "EPS Grw": False}
    try:
        info = yf.Ticker(yf_ticker).info
        if not info:
            return result, is_ttm

        result["Mkt Cap"] = info.get("marketCap")
        result["Fwd P/E"] = info.get("forwardPE")
        result["EV/EBITDA"] = info.get("enterpriseToEbitda")

        # Margins from yfinance are 0-1 decimals, convert to percentage
        gm = info.get("grossMargins")
        result["Gross Mgn"] = gm * 100 if gm is not None else None
        om = info.get("operatingMargins")
        result["Op Mgn"] = om * 100 if om is not None else None
        roe = info.get("returnOnEquity")
        result["ROE"] = roe * 100 if roe is not None else None

        # Growth — yfinance gives quarterly YoY as fallback
        rev_grw = info.get("revenueGrowth")
        result["Rev Grw"] = rev_grw * 100 if rev_grw is not None else None
        eps_grw = info.get("earningsGrowth")
        result["EPS Grw"] = eps_grw * 100 if eps_grw is not None else None

        # PEG from yfinance
        result["PEG"] = info.get("pegRatio")
    except Exception:
        pass

    # Enrich with Finnhub TTM YoY data (US tickers only, more accurate for growth)
    if finnhub_metrics:
        fh_rev = finnhub_metrics.get("revenueGrowthTTMYoy")
        if fh_rev is not None:
            result["Rev Grw"] = fh_rev
            is_ttm["Rev Grw"] = True
        fh_eps = finnhub_metrics.get("epsGrowthTTMYoy")
        if fh_eps is not None:
            result["EPS Grw"] = fh_eps
            is_ttm["EPS Grw"] = True
        fh_peg = finnhub_metrics.get("pegTTM")
        if fh_peg is not None:
            result["PEG"] = fh_peg

    return result, is_ttm


def format_mkt_cap(value):
    """Format market cap as abbreviated string (e.g., $26.2B)."""
    if value is None or pd.isna(value):
        return "N/A"
    v = float(value)
    if v >= 1e12:
        return f"${v/1e12:.1f}T"
    if v >= 1e9:
        return f"${v/1e9:.1f}B"
    if v >= 1e6:
        return f"${v/1e6:.0f}M"
    return f"${v:.0f}"


def calc_annual_return(hist, year):
    """Calculate total return for a given calendar year."""
    year_data = hist[hist.index.year == year]
    if len(year_data) < 2:
        return None
    first = year_data.iloc[0]
    last = year_data.iloc[-1]
    return (last / first - 1) * 100


def calc_period_return(hist, days):
    """Calculate return over last N calendar days from latest data point."""
    if hist.empty:
        return None
    end = hist.iloc[-1]
    target_date = hist.index[-1] - pd.Timedelta(days=days)
    earlier = hist[hist.index <= target_date]
    if earlier.empty:
        return None
    start = earlier.iloc[-1]
    return (end / start - 1) * 100


def calc_1d_return(hist):
    """Calculate 1-day return (last close vs prior close)."""
    if len(hist) < 2:
        return None
    return (hist.iloc[-1] / hist.iloc[-2] - 1) * 100


def calc_1w_return(hist):
    """Calculate 1-week return (last 5 trading days)."""
    if len(hist) < 6:
        return None
    return (hist.iloc[-1] / hist.iloc[-6] - 1) * 100


def calc_qtd_return(hist):
    """Calculate quarter-to-date return."""
    today = date.today()
    # Quarter start month: Jan=1, Apr=4, Jul=7, Oct=10
    q_start_month = ((today.month - 1) // 3) * 3 + 1
    q_start = pd.Timestamp(today.year, q_start_month, 1)
    qtr_data = hist[hist.index >= q_start]
    if len(qtr_data) < 2:
        return None
    return (qtr_data.iloc[-1] / qtr_data.iloc[0] - 1) * 100


def calc_ytd_return(hist):
    """Calculate YTD return."""
    today = date.today()
    year_data = hist[hist.index.year == today.year]
    if len(year_data) < 2:
        return None
    first = year_data.iloc[0]
    last = year_data.iloc[-1]
    return (last / first - 1) * 100


def get_color(value, is_negative_col=False):
    """Return RGB hex color for a return value. Red for negative, green for positive, white for zero/NA."""
    if value is None or pd.isna(value):
        return "FFFFFF"
    val = float(value)
    if val == 0:
        return "FFFFFF"
    if val > 0:
        # Green scale: 0% -> white, 100%+ -> deep green
        intensity = min(abs(val) / 100, 1.0)
        r = int(255 - intensity * 155)  # 255 -> 100
        g = int(255 - intensity * 30)   # 255 -> 225
        b = int(255 - intensity * 155)  # 255 -> 100
        return f"{r:02X}{g:02X}{b:02X}"
    else:
        # Red scale: 0% -> white, -100%+ -> deep red
        intensity = min(abs(val) / 100, 1.0)
        r = int(255 - intensity * 30)   # 255 -> 225
        g = int(255 - intensity * 155)  # 255 -> 100
        b = int(255 - intensity * 155)  # 255 -> 100
        return f"{r:02X}{g:02X}{b:02X}"


def get_html_color(value):
    """Return CSS background-color for HTML."""
    hex_color = get_color(value)
    return f"#{hex_color}"


def build_ticker_health_data(df_unique, yf_tickers, ticker_map, all_results, all_fundamentals):
    """Build health report data from already-collected information (no extra API calls)."""
    # Tickers with no price data
    no_price = [ticker_map[t] for t in yf_tickers if t not in all_results]

    # Tickers with no fundamental data (all values None)
    no_fundamentals = []
    for yf_t in yf_tickers:
        fund = all_fundamentals.get(yf_t, {})
        if all(v is None for v in fund.values()):
            no_fundamentals.append(ticker_map.get(yf_t, yf_t))

    # Missing metadata from CSV
    missing_exchange = []
    missing_company_name = []
    for _, row in df_unique.iterrows():
        ticker = str(row.get("Ticker", "")).strip()
        if not ticker or ticker == "#N/A":
            continue
        exchange = str(row.get("Exchange", "")).strip()
        if not exchange or exchange == "nan":
            missing_exchange.append(ticker)
        company = str(row.get("Company Name", "")).strip()
        if not company or company == "nan":
            missing_company_name.append(ticker)

    total = len(no_price) + len(no_fundamentals) + len(missing_exchange) + len(missing_company_name)
    return {
        "no_price": sorted(no_price),
        "no_fundamentals": sorted(no_fundamentals),
        "missing_exchange": sorted(missing_exchange),
        "missing_company_name": sorted(missing_company_name),
        "total_issues": total,
    }


def generate_health_html(health_data):
    """Generate collapsible HTML section for ticker health report."""
    total = health_data["total_issues"]
    if total == 0:
        badge_class = "health-badge ok"
        badge_text = "All clear"
    elif health_data["no_price"]:
        badge_class = "health-badge critical"
        badge_text = f"{total} issues"
    else:
        badge_class = "health-badge warning"
        badge_text = f"{total} issues"

    def ticker_list_html(tickers, max_show=50):
        if not tickers:
            return "<p>None</p>"
        items = ", ".join(tickers[:max_show])
        extra = f" ... and {len(tickers) - max_show} more" if len(tickers) > max_show else ""
        return f"<p>{items}{extra}</p>"

    sections = []

    if health_data["no_price"]:
        sections.append(f'''
  <div class="health-section critical">
    <h4>No Price Data — Potential Delistings ({len(health_data["no_price"])})</h4>
    {ticker_list_html(health_data["no_price"])}
  </div>''')

    if health_data["no_fundamentals"]:
        sections.append(f'''
  <div class="health-section warning">
    <h4>No Fundamental Data ({len(health_data["no_fundamentals"])})</h4>
    {ticker_list_html(health_data["no_fundamentals"])}
  </div>''')

    if health_data["missing_exchange"]:
        sections.append(f'''
  <div class="health-section info">
    <h4>Missing Exchange ({len(health_data["missing_exchange"])})</h4>
    {ticker_list_html(health_data["missing_exchange"])}
  </div>''')

    if health_data["missing_company_name"]:
        sections.append(f'''
  <div class="health-section info">
    <h4>Missing Company Name ({len(health_data["missing_company_name"])})</h4>
    {ticker_list_html(health_data["missing_company_name"])}
  </div>''')

    return f'''<details class="ticker-health">
  <summary>Ticker Health Report <span class="{badge_class}">{badge_text}</span></summary>
{"".join(sections)}
</details>'''


def archive_old_files():
    """Move prior dated performance files to old reports folder."""
    os.makedirs(OLD_REPORTS_DIR, exist_ok=True)
    patterns = [
        os.path.join(REPORTS_DIR, "coverage_performance_*.xlsx"),
        os.path.join(REPORTS_DIR, "coverage_performance_*.html"),
    ]
    moved = 0
    for pattern in patterns:
        for f in glob.glob(pattern):
            # Don't move today's files
            if TODAY in os.path.basename(f):
                continue
            dest = os.path.join(OLD_REPORTS_DIR, os.path.basename(f))
            shutil.move(f, dest)
            moved += 1
    if moved:
        print(f"Archived {moved} old file(s) to: {OLD_REPORTS_DIR}")


def send_email_report(gmail_addr, gmail_pass, html_path):
    """Send the HTML report as an email attachment via Gmail SMTP."""
    msg = MIMEMultipart()
    msg["From"] = gmail_addr
    msg["To"] = gmail_addr
    msg["Subject"] = f"Coverage Performance Report — {TODAY}"

    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")
    msg.attach(MIMEText(f"Report attached, generated {timestamp}.", "plain"))

    with open(html_path, "rb") as f:
        part = MIMEBase("text", "html")
        part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header(
            "Content-Disposition",
            f"attachment; filename={os.path.basename(html_path)}",
        )
        msg.attach(part)

    with smtplib.SMTP("smtp.gmail.com", 587) as server:
        server.starttls()
        server.login(gmail_addr, gmail_pass)
        server.send_message(msg)
    print(f"Emailed report to {gmail_addr}")


def main():
    os.makedirs(REPORTS_DIR, exist_ok=True)
    archive_old_files()
    print("Reading coverage CSV...")
    df = pd.read_csv(CSV_PATH)

    # Deduplicate tickers
    seen = set()
    unique_rows = []
    for _, row in df.iterrows():
        ticker = str(row.get("Ticker", "")).strip()
        if ticker and ticker != "#N/A" and ticker not in seen:
            seen.add(ticker)
            unique_rows.append(row)
    df_unique = pd.DataFrame(unique_rows).reset_index(drop=True)
    print(f"Found {len(df_unique)} unique tickers")

    # Normalize tickers for yfinance
    yf_tickers = []
    ticker_map = {}  # yf_ticker -> original_ticker
    for _, row in df_unique.iterrows():
        orig = str(row["Ticker"]).strip()
        company = str(row.get("Company Name", "")).strip()
        yf_t = normalize_ticker(orig, company)
        if yf_t:
            yf_tickers.append(yf_t)
            ticker_map[yf_t] = orig

    print(f"Downloading data for {len(yf_tickers)} tickers (this may take several minutes)...")

    # Download in batches to avoid timeouts
    batch_size = 50
    all_results = {}
    total_batches = (len(yf_tickers) + batch_size - 1) // batch_size

    for i in range(0, len(yf_tickers), batch_size):
        batch = yf_tickers[i:i + batch_size]
        batch_num = i // batch_size + 1
        print(f"  Batch {batch_num}/{total_batches} ({len(batch)} tickers)...")
        try:
            data = yf.download(
                batch,
                start="2015-01-01",
                auto_adjust=True,
                progress=False,
                threads=True,
            )
            if data.empty:
                continue
            close = data["Close"]
            if isinstance(close, pd.Series):
                # Single ticker returned as Series
                t = batch[0]
                all_results[t] = close.dropna()
            else:
                for t in batch:
                    if t in close.columns:
                        series = close[t].dropna()
                        if len(series) > 0:
                            all_results[t] = series
        except Exception as e:
            print(f"  Error in batch {batch_num}: {e}")
            continue

    print(f"Successfully downloaded data for {len(all_results)} tickers")

    # FMP fallback for missing US tickers
    fmp_key = API_KEYS.get("FMP_API_KEY")
    if fmp_key:
        missing = [t for t in yf_tickers if t not in all_results and "." not in t]
        if missing:
            print(f"Trying FMP API fallback for {len(missing)} missing US tickers...")
            fmp_found = 0
            for t in missing:
                series = try_fmp_historical(t, fmp_key)
                if series is not None:
                    all_results[t] = series
                    fmp_found += 1
            print(f"  FMP resolved {fmp_found} additional tickers")

    print(f"Total tickers with data: {len(all_results)}")

    # Fetch Finnhub metrics for US tickers (no dot in symbol = likely US)
    finnhub_key = API_KEYS.get("FINNHUB_API_KEY")
    finnhub_data = {}
    if finnhub_key:
        us_yf_tickers = [t for t in yf_tickers if "." not in t]
        print(f"Fetching Finnhub fundamentals for {len(us_yf_tickers)} US tickers...")
        import time
        for i, t in enumerate(us_yf_tickers):
            if i > 0 and i % 60 == 0:
                print(f"  {i}/{len(us_yf_tickers)} (pausing for rate limit)...")
                time.sleep(1)
            metrics = fetch_finnhub_metrics(t, finnhub_key)
            if metrics:
                finnhub_data[t] = metrics
        print(f"  Finnhub returned data for {len(finnhub_data)} tickers")

    # Fetch yfinance fundamentals for all tickers
    print(f"Fetching yfinance fundamentals for {len(yf_tickers)} tickers...")
    all_fundamentals = {}
    all_is_ttm = {}  # yf_ticker -> {"Rev Grw": bool, "EPS Grw": bool}
    for i, yf_t in enumerate(yf_tickers):
        if i > 0 and i % 100 == 0:
            print(f"  {i}/{len(yf_tickers)}...")
        fh = finnhub_data.get(yf_t)
        fund_result, is_ttm = fetch_fundamentals(yf_t, finnhub_metrics=fh)
        all_fundamentals[yf_t] = fund_result
        all_is_ttm[yf_t] = is_ttm
    fund_count = sum(1 for v in all_fundamentals.values() if v.get("Mkt Cap") is not None)
    print(f"  Fundamentals loaded for {fund_count} tickers")

    # Build ticker health data (no extra API calls)
    health_data = build_ticker_health_data(df_unique, yf_tickers, ticker_map, all_results, all_fundamentals)
    print(f"Ticker health: {health_data['total_issues']} issues found")

    # Calculate returns
    results = []
    for _, row in df_unique.iterrows():
        orig_ticker = str(row["Ticker"]).strip()
        company = str(row.get("Company Name", "")).strip()
        yf_t = normalize_ticker(orig_ticker, company)
        sector = str(row.get("Sector (JP)", row.get("Sector", ""))).strip()
        subsector = str(row.get("Subsector (JP)", row.get("Subsector", ""))).strip()

        returns = {}
        if yf_t and yf_t in all_results:
            hist = all_results[yf_t]
            for year in ANNUAL_YEARS:
                returns[str(year)] = calc_annual_return(hist, year)
            returns["1D"] = calc_1d_return(hist)
            returns["1W"] = calc_1w_return(hist)
            returns["QTD"] = calc_qtd_return(hist)
            returns["YTD"] = calc_ytd_return(hist)
            returns["1Y"] = calc_period_return(hist, 365)
            returns["3Y"] = calc_period_return(hist, 365 * 3)
            returns["5Y"] = calc_period_return(hist, 365 * 5)
            returns["10Y"] = calc_period_return(hist, 365 * 10)
        else:
            for col in RETURN_COLS:
                returns[col] = None

        # Fundamentals
        fund = all_fundamentals.get(yf_t, {col: None for col in FUND_COLS})
        is_ttm = all_is_ttm.get(yf_t, {"Rev Grw": False, "EPS Grw": False})

        result_row = {
            "Ticker": orig_ticker,
            "Company Name": company if company != "nan" else "",
            "Sector (JP)": sector if sector != "nan" else "",
            "Subsector (JP)": subsector if subsector != "nan" else "",
            "_is_ttm_rev": is_ttm["Rev Grw"],
            "_is_ttm_eps": is_ttm["EPS Grw"],
        }
        result_row.update(returns)
        result_row.update(fund)
        results.append(result_row)

    result_df = pd.DataFrame(results)
    info_cols = ["Ticker", "Company Name", "Sector (JP)", "Subsector (JP)"]

    # ============ EXCEL OUTPUT ============
    print("Generating Excel file...")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Coverage Performance"

    # Header style
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )

    all_cols = info_cols + RETURN_COLS + FUND_COLS
    # Write headers
    fund_header_fill = PatternFill(start_color="1A5276", end_color="1A5276", fill_type="solid")
    for col_idx, col_name in enumerate(all_cols, 1):
        display_name = FUND_DISPLAY_NAMES.get(col_name, col_name)
        cell = ws.cell(row=1, column=col_idx, value=display_name)
        cell.fill = fund_header_fill if col_name in FUND_COLS else header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border

    # Write data
    for row_idx, (_, row) in enumerate(result_df.iterrows(), 2):
        for col_idx, col_name in enumerate(all_cols, 1):
            val = row.get(col_name)
            if col_name in RETURN_COLS:
                if val is not None and not pd.isna(val):
                    cell = ws.cell(row=row_idx, column=col_idx, value=round(val, 1))
                    cell.number_format = '0.0"%"'
                    hex_color = get_color(val)
                    cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
                    if val < 0:
                        cell.font = Font(color="8B0000", size=9)
                    elif val > 0:
                        cell.font = Font(color="006400", size=9)
                    else:
                        cell.font = Font(size=9)
                else:
                    cell = ws.cell(row=row_idx, column=col_idx, value="N/A")
                    cell.font = Font(color="999999", size=9)
            elif col_name in FUND_COLS:
                # Check if growth value needs asterisk (non-TTM source)
                needs_asterisk = False
                if col_name == "Rev Grw" and not row.get("_is_ttm_rev", False):
                    needs_asterisk = True
                elif col_name == "EPS Grw" and not row.get("_is_ttm_eps", False):
                    needs_asterisk = True
                if col_name in FUND_MONEY_COLS:
                    cell = ws.cell(row=row_idx, column=col_idx, value=format_mkt_cap(val))
                    cell.font = Font(size=9)
                elif val is not None and not pd.isna(val):
                    if col_name in FUND_PCT_COLS:
                        asterisk = "*" if needs_asterisk else ""
                        cell = ws.cell(row=row_idx, column=col_idx, value=f"{val:.1f}%{asterisk}")
                        cell.number_format = '@'
                        hex_color = get_color(val)
                        cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
                        if val < 0:
                            cell.font = Font(color="8B0000", size=9)
                        elif val > 0:
                            cell.font = Font(color="006400", size=9)
                        else:
                            cell.font = Font(size=9)
                    else:
                        cell = ws.cell(row=row_idx, column=col_idx, value=round(val, 1))
                        cell.number_format = '0.0'
                        cell.font = Font(size=9)
                else:
                    cell = ws.cell(row=row_idx, column=col_idx, value="N/A")
                    cell.font = Font(color="999999", size=9)
            else:
                display_val = val if val and str(val) != "nan" else ""
                cell = ws.cell(row=row_idx, column=col_idx, value=display_val)
                cell.font = Font(size=9)

            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center" if col_name in RETURN_COLS or col_name in FUND_COLS else "left", vertical="center")

    # Column widths
    ws.column_dimensions["A"].width = 12  # Ticker
    ws.column_dimensions["B"].width = 30  # Company Name
    ws.column_dimensions["C"].width = 18  # Sector (JP)
    ws.column_dimensions["D"].width = 22  # Subsector (JP)
    for col_idx in range(5, 5 + len(RETURN_COLS)):
        ws.column_dimensions[get_column_letter(col_idx)].width = 10
    fund_start = 5 + len(RETURN_COLS)
    for i, col_name in enumerate(FUND_COLS):
        col_letter = get_column_letter(fund_start + i)
        ws.column_dimensions[col_letter].width = 12 if col_name == "Mkt Cap" else 10

    # Freeze panes
    ws.freeze_panes = "E2"

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(all_cols))}{len(result_df) + 1}"

    # Add footnote row for asterisk explanation
    footnote_row = len(result_df) + 3
    cell = ws.cell(row=footnote_row, column=1,
                   value="* Value reflects quarterly YoY growth (yfinance) rather than TTM YoY (Finnhub). TTM YoY data was not available for this ticker.")
    cell.font = Font(size=8, italic=True, color="888888")
    ws.merge_cells(start_row=footnote_row, start_column=1, end_row=footnote_row, end_column=8)

    wb.save(OUTPUT_XLSX)
    print(f"Saved: {OUTPUT_XLSX}")

    # ============ HTML OUTPUT ============
    print("Generating HTML file...")

    # HTML column order: Ticker, Mkt Cap, Company Name, Sector, Subsector, [returns], [fundamentals minus Mkt Cap]
    html_info_cols = ["Ticker", "Mkt Cap", "Company Name", "Sector (JP)", "Subsector (JP)"]
    html_fund_cols = [c for c in FUND_COLS if c != "Mkt Cap"]

    # Collect unique sectors for filter dropdown
    all_sectors = sorted(set(
        str(row.get("Sector (JP)", "")).strip()
        for _, row in result_df.iterrows()
        if str(row.get("Sector (JP)", "")).strip() and str(row.get("Sector (JP)", "")).strip() != "nan"
    ))

    html_rows = []
    for _, row in result_df.iterrows():
        cells = []
        # Ticker
        val = row.get("Ticker", "")
        if pd.isna(val) or str(val) == "nan":
            val = ""
        cells.append(f'<td class="info">{val}</td>')
        # Mkt Cap (moved here)
        mcap = row.get("Mkt Cap")
        cells.append(f'<td class="fund">{format_mkt_cap(mcap)}</td>')
        # Company Name, Sector, Subsector
        for col in ["Company Name", "Sector (JP)", "Subsector (JP)"]:
            val = row.get(col, "")
            if pd.isna(val) or str(val) == "nan":
                val = ""
            cells.append(f'<td class="info">{val}</td>')
        # Return columns
        for col in RETURN_COLS:
            val = row.get(col)
            if val is not None and not pd.isna(val):
                bg = get_html_color(val)
                text_color = "#8B0000" if val < 0 else "#006400" if val > 0 else "#333"
                cells.append(f'<td class="ret" style="background-color:{bg};color:{text_color}">{val:.1f}%</td>')
            else:
                cells.append('<td class="ret na">N/A</td>')
        # Fundamental columns (minus Mkt Cap)
        for col in html_fund_cols:
            val = row.get(col)
            # Check if growth value needs asterisk (non-TTM source)
            needs_asterisk = False
            if col == "Rev Grw" and not row.get("_is_ttm_rev", False):
                needs_asterisk = True
            elif col == "EPS Grw" and not row.get("_is_ttm_eps", False):
                needs_asterisk = True
            if val is not None and not pd.isna(val):
                if col in FUND_PCT_COLS:
                    asterisk = "*" if needs_asterisk else ""
                    bg = get_html_color(val)
                    text_color = "#8B0000" if val < 0 else "#006400" if val > 0 else "#333"
                    cells.append(f'<td class="fund" style="background-color:{bg};color:{text_color}">{val:.1f}%{asterisk}</td>')
                else:
                    cells.append(f'<td class="fund">{val:.1f}x</td>')
            else:
                cells.append('<td class="fund na">N/A</td>')
        html_rows.append("<tr>" + "".join(cells) + "</tr>")

    html_info_display = [FUND_DISPLAY_NAMES.get(c, c) for c in html_info_cols]
    info_headers = "".join(f"<th>{c}</th>" for c in html_info_display)
    ret_headers = "".join(f"<th>{c}</th>" for c in RETURN_COLS)
    fund_headers = "".join(f'<th class="fund-hdr">{FUND_DISPLAY_NAMES.get(c, c)}</th>' for c in html_fund_cols)
    header_cells = info_headers + ret_headers + fund_headers
    sector_options = "".join(f'<option value="{s}">{s}</option>' for s in all_sectors)
    timestamp = datetime.now().strftime("%B %d, %Y %H:%M")

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>Coverage Universe Performance</title>
<style>
  body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Arial, sans-serif; margin: 20px; background: #f5f5f5; }}
  h1 {{ color: #2c3e50; font-size: 22px; }}
  p.timestamp {{ color: #888; font-size: 12px; margin-bottom: 15px; }}
  .table-wrapper {{ overflow-x: auto; max-height: 85vh; overflow-y: auto; }}
  table {{ border-collapse: collapse; font-size: 11px; width: 100%; }}
  thead {{ position: sticky; top: 0; z-index: 2; }}
  th {{ background: #2c3e50; color: white; padding: 8px 6px; text-align: center; border: 1px solid #1a252f;
       font-weight: 600; white-space: nowrap; cursor: pointer; user-select: none; }}
  th:hover {{ background: #3e5871; }}
  th.sort-asc::after {{ content: " ▲"; font-size: 9px; }}
  th.sort-desc::after {{ content: " ▼"; font-size: 9px; }}
  td {{ padding: 5px 6px; border: 1px solid #ddd; white-space: nowrap; }}
  td.info {{ text-align: left; background: #fff; }}
  td.ret {{ text-align: center; font-weight: 500; }}
  td.fund {{ text-align: center; font-weight: 500; }}
  td.na {{ color: #bbb; background: #fafafa; }}
  th.fund-hdr {{ background: #1a5276; }}
  tr:hover td {{ opacity: 0.85; }}
  tr:nth-child(even) td.info {{ background: #f9f9f9; }}
  .filter-bar {{ margin-bottom: 12px; display: flex; align-items: center; gap: 10px; }}
  .filter-bar label {{ font-size: 13px; font-weight: 600; color: #2c3e50; }}
  .filter-bar select {{ font-size: 13px; padding: 4px 8px; border: 1px solid #ccc; border-radius: 4px; }}
  .filter-bar .count {{ font-size: 12px; color: #888; }}
  .footnote {{ font-size: 11px; color: #888; margin-top: 8px; font-style: italic; }}
  .ticker-health {{ margin-top: 20px; background: #fff; border: 1px solid #ddd; border-radius: 6px; padding: 10px 15px; }}
  .ticker-health summary {{ font-size: 14px; font-weight: 600; color: #2c3e50; cursor: pointer; }}
  .health-badge {{ display: inline-block; padding: 2px 8px; border-radius: 10px; font-size: 11px; font-weight: 600; margin-left: 8px; }}
  .health-badge.ok {{ background: #d4edda; color: #155724; }}
  .health-badge.warning {{ background: #fff3cd; color: #856404; }}
  .health-badge.critical {{ background: #f8d7da; color: #721c24; }}
  .health-section {{ margin: 10px 0; padding: 8px 12px; border-radius: 4px; }}
  .health-section h4 {{ margin: 0 0 4px 0; font-size: 12px; }}
  .health-section p {{ margin: 4px 0; font-size: 11px; word-break: break-all; }}
  .health-section.critical {{ background: #f8d7da; border-left: 3px solid #dc3545; }}
  .health-section.warning {{ background: #fff3cd; border-left: 3px solid #ffc107; }}
  .health-section.info {{ background: #f0f0f0; border-left: 3px solid #999; }}
</style>
</head>
<body>
<h1>Coverage Universe Performance</h1>
<p class="timestamp">Generated: {timestamp}</p>
<p class="footnote">* Value reflects quarterly YoY growth (yfinance) rather than TTM YoY (Finnhub). TTM YoY data was not available for this ticker.</p>
<div class="filter-bar">
  <label for="sectorFilter">Sector:</label>
  <select id="sectorFilter">
    <option value="">All Sectors</option>
    {sector_options}
  </select>
  <span class="count" id="rowCount"></span>
</div>
<div class="table-wrapper">
<table>
<thead><tr>{header_cells}</tr></thead>
<tbody>
{"".join(html_rows)}
</tbody>
</table>
</div>
<script>
var SECTOR_COL = 3;
function updateCount() {{
  var tbody = document.querySelector("tbody");
  var visible = tbody.querySelectorAll("tr:not([style*='display: none'])").length;
  var total = tbody.querySelectorAll("tr").length;
  document.getElementById("rowCount").textContent = visible + " of " + total + " companies";
}}
// Sector filter
document.getElementById("sectorFilter").addEventListener("change", function() {{
  var val = this.value.toLowerCase();
  document.querySelectorAll("tbody tr").forEach(function(row) {{
    var sector = row.children[SECTOR_COL].textContent.trim().toLowerCase();
    row.style.display = (!val || sector.includes(val)) ? "" : "none";
  }});
  updateCount();
}});
// Column sorting
document.querySelectorAll("th").forEach(function(th, colIdx) {{
  th.addEventListener("click", function() {{
    var table = th.closest("table");
    var tbody = table.querySelector("tbody");
    var rows = Array.from(tbody.querySelectorAll("tr"));
    var asc = !th.classList.contains("sort-asc");
    table.querySelectorAll("th").forEach(function(h) {{ h.classList.remove("sort-asc","sort-desc"); }});
    th.classList.add(asc ? "sort-asc" : "sort-desc");
    rows.sort(function(a, b) {{
      var aText = a.children[colIdx].textContent.trim();
      var bText = b.children[colIdx].textContent.trim();
      var aNum = parseFloat(aText.replace(/[$%xTBMNA,]/g, ""));
      var bNum = parseFloat(bText.replace(/[$%xTBMNA,]/g, ""));
      if (aText.includes("T")) aNum *= 1e12; else if (aText.includes("B")) aNum *= 1e9; else if (aText.match(/\$.*M/)) aNum *= 1e6;
      if (bText.includes("T")) bNum *= 1e12; else if (bText.includes("B")) bNum *= 1e9; else if (bText.match(/\$.*M/)) bNum *= 1e6;
      var aNA = isNaN(aNum); var bNA = isNaN(bNum);
      if (aNA && bNA) return aText.localeCompare(bText) * (asc ? 1 : -1);
      if (aNA) return 1;
      if (bNA) return -1;
      return asc ? aNum - bNum : bNum - aNum;
    }});
    rows.forEach(function(r) {{ tbody.appendChild(r); }});
  }});
}});
updateCount();
</script>
{generate_health_html(health_data)}
</body>
</html>"""

    with open(OUTPUT_HTML, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"Saved: {OUTPUT_HTML}")

    # ============ EMAIL REPORT ============
    gmail_addr = API_KEYS.get("GMAIL_ADDRESS")
    gmail_pass = API_KEYS.get("GMAIL_APP_PASSWORD")
    if gmail_addr and gmail_pass:
        print("Emailing HTML report...")
        send_email_report(gmail_addr, gmail_pass, OUTPUT_HTML)
    else:
        print("Skipping email (GMAIL_ADDRESS / GMAIL_APP_PASSWORD not set in .env)")

    print("Done!")


if __name__ == "__main__":
    main()
