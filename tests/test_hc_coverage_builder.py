"""`scripts/build_hc_coverage_xlsx.py` — the AA_Core Coverage workbook.

Codex reviewed the design on 2026-08-26 and its last finding was that the builder
had no tests at all, which is how a `.info`-per-field fetch shipped in 3be1f05,
throttled Yahoo for an entire session, and was found by review rather than by CI.
Every test here pins a defect that actually occurred or a rule that would publish
something private if it silently stopped holding.
"""
import datetime
import importlib.util
import os

import openpyxl
import pytest

_SPEC = importlib.util.spec_from_file_location(
    "hc_builder",
    os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                 "scripts", "build_hc_coverage_xlsx.py"))
b = importlib.util.module_from_spec(_SPEC)
_SPEC.loader.exec_module(b)


# ── the regression that caused the outage ────────────────────────────────────

def test_info_is_fetched_once_per_attempt_not_once_per_field(monkeypatch):
    """THE bug. It was written as:

        d = {k: (yf.Ticker(sym).info or {}).get(k) for k in FIELDS}

    which constructs a Ticker and hits `.info` once for EVERY field. Six fields x
    four retry attempts x 239 rows is up to 5,736 requests where 239 would do. It
    throttled Yahoo hard enough that three consecutive builds were refused 90,
    136 and 143 rows respectively and every one fell through to FMP. After the
    fix the next build took all 239 from Yahoo with zero fallbacks.

    One row, one healthy payload => exactly ONE `.info` access.
    """
    hits = []

    class FakeTicker:
        def __init__(self, sym):
            self.sym = sym

        @property
        def info(self):
            hits.append(self.sym)
            return {"marketCap": 1e9, "regularMarketPrice": 10.0,
                    "currency": "USD", "longName": "Acme Inc"}

    monkeypatch.setitem(__import__("sys").modules, "yfinance",
                        type("m", (), {"Ticker": FakeTicker}))
    out = b.fetch([{"Ticker": "ACME", "Company Name": "Acme Inc",
                    "Exchange": "NASDAQ"}])
    assert out["ACME"]["marketCap"] == 1e9
    assert len(hits) == 1, (
        "`.info` was accessed %d times for ONE ticker. It is being re-fetched per "
        "field again, which is what throttled Yahoo on 2026-08-26." % len(hits))


# ── what may and may not be published ────────────────────────────────────────

def test_rating_sits_immediately_after_company_name():
    """JP 2026-08-26: "have a ratings column after the company name column"."""
    assert b.COLS[:3] == ["Ticker", "Company Name", "Rating"]


def test_ramp_effort_is_gone_from_every_surface():
    assert not any("Ramp" in c for c in b.COLS)
    assert not any("Ramp" in c for c in b.PUBLIC_COLS)


def test_the_private_only_mechanism_still_works_even_though_it_is_empty():
    """`Rating` was withheld from the public CSV until JP asked for it in the
    Google file too, so PRIVATE_ONLY is empty now. Keep the mechanism honest: the
    next sensitive column must be excludable without rediscovering that the CSV
    writer publishes everything in COLS."""
    assert b.PUBLIC_COLS == [c for c in b.COLS if c not in b.PRIVATE_ONLY]
    probe = set(b.COLS[:1])
    assert [c for c in b.COLS if c not in probe] == b.COLS[1:]


def test_public_schema_is_the_private_one_minus_exactly_the_private_columns():
    """Pins the relationship rather than a column count, so a new column is
    published deliberately or not at all."""
    assert set(b.PUBLIC_COLS) == set(b.COLS) - b.PRIVATE_ONLY
    assert b.PUBLIC_COLS == [c for c in b.COLS if c not in b.PRIVATE_ONLY], \
        "public column ORDER must track COLS; the Sheet reads by position"


def test_the_published_csv_on_disk_carries_no_rating_column():
    """Belt and braces: check the artifact, not just the constant. A published
    artifact is the thing consumers read, and validating the source instead of the
    artifact is how a BOM once emptied every export."""
    path = b.PUBLIC_CSV
    if not os.path.exists(path):
        pytest.skip("public CSV not built in this checkout")
    lines = open(path, encoding="utf-8").read().splitlines()
    # Row 1 is the provenance line, row 2 blank, row 3 the header. JP asked for
    # "when it was last updated and any relevant background" to live in the file,
    # and the Google Sheet is a single =IMPORTDATA cell that nothing here can write
    # to -- so the only way it reaches that surface is inside the CSV itself.
    assert "LAST UPDATED" in lines[0], "the published CSV lost its provenance line"
    header = lines[2].split(",")
    assert header[:4] == ["#", "Ticker", "Company Name", "Rating"], header[:4]
    assert "Size" in header and "Fwd P/E" in header
    assert not any("Ramp" in h for h in header)


# ── the size bucket ──────────────────────────────────────────────────────────

@pytest.mark.parametrize("mcap, expected", [
    (22613.4, "SMID"),   # Guardant — SMID on JP's reference sheet
    (34126.0, "LC"),     # Illumina — LC on JP's reference sheet
    (b.LC_THRESHOLD_USD_M, "LC"),        # boundary is inclusive
    (b.LC_THRESHOLD_USD_M - 0.01, "SMID"),
])
def test_size_bucket(mcap, expected):
    assert b.size_bucket(mcap) == expected


def test_size_is_blank_when_market_cap_is_unknown():
    """Never a guessed bucket. The partial-book guard tolerates up to 5% of rows
    missing a market cap, so this is reachable, and the proposal that claimed it
    "cannot happen" was wrong."""
    assert b.size_bucket(None) is None


# ── forward P/E ──────────────────────────────────────────────────────────────

def test_forward_pe_passes_through_a_real_multiple():
    assert b.forward_pe(22.1) == 22.1


@pytest.mark.parametrize("raw", [-380.9, 0, None, "", "n/a"])
def test_forward_pe_blanks_anything_that_is_not_a_positive_multiple(raw):
    """Yahoo returns a NEGATIVE forwardPE for a company expected to lose money —
    Guardant came back -380.9. Sorting a column containing it puts the biggest
    loss-maker at the top as though it were the cheapest name on the sheet."""
    assert b.forward_pe(raw) is None


# ── the ratings join ─────────────────────────────────────────────────────────

def _ratings_book(tmp_path, rows):
    path = tmp_path / "Ratings_CoreCoverage.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = b.RATING_SHEET
    for j, c in enumerate(b.RATING_COLS, start=1):
        ws.cell(1, j, c)
    for i, r in enumerate(rows, start=2):
        for j, c in enumerate(b.RATING_COLS, start=1):
            ws.cell(i, j, r.get(c))
    wb.save(path)
    return str(path)


def test_a_rating_loads_for_a_matching_row(tmp_path, monkeypatch):
    monkeypatch.setattr(b, "RATINGS_PATH", _ratings_book(tmp_path, [
        {"Ticker": "ISRG", "Company Name": "Intuitive Surgical Inc", "Rating": "2"},
    ]))
    loaded = b.load_ratings()
    assert loaded["ISRG"]["Rating"] == "2"


def test_duplicate_tickers_in_the_ratings_file_abort_rather_than_pick_one(
        tmp_path, monkeypatch):
    """A duplicated join key fans a left-join out into extra rows, and silently
    choosing one of two ratings is choosing for JP."""
    monkeypatch.setattr(b, "RATINGS_PATH", _ratings_book(tmp_path, [
        {"Ticker": "ISRG", "Company Name": "Intuitive Surgical Inc", "Rating": "2"},
        {"Ticker": "ISRG", "Company Name": "Intuitive Surgical Inc", "Rating": "4"},
    ]))
    with pytest.raises(SystemExit) as e:
        b.load_ratings()
    assert "duplicate" in str(e.value).lower()


def test_a_missing_ratings_file_is_blank_not_an_error(tmp_path, monkeypatch):
    monkeypatch.setattr(b, "RATINGS_PATH", str(tmp_path / "nope.xlsx"))
    assert b.load_ratings() == {}


def test_a_rating_does_not_attach_when_the_ratings_file_names_another_company():
    """JP: "Ticker is an identity but it can be fuzzy so you need to check and
    verify with me if its too ambiguous."

    So the rating must become UNREACHABLE, not merely warned about. `ZEN` is the
    precedent: the ticker moved from Zendesk to Zentek and the stale
    classification rode along because nothing refused to use it."""
    assert b._payload_names_match("Zentek Ltd", "ZEN TECHNOLOGIES LTD") is False
    assert b._payload_names_match("Medartis Holding AG", "Medifast, Inc.") is False
    # ...while the same company spelled two ways still joins.
    assert b._payload_names_match("bioMerieux SA", "bioMérieux S.A.") is True


def test_a_blank_name_in_the_ratings_file_still_joins():
    """If JP clears a name cell the rating should still attach — an absent name is
    not evidence of a mismatch, and the alternative silently drops his work."""
    assert b._payload_names_match("", "Anything Inc") is True


# ── returns freshness ────────────────────────────────────────────────────────

def test_a_stale_snapshot_yields_blank_returns_not_stale_ones(monkeypatch, tmp_path):
    """A YTD from a month ago sitting beside a same-day price reads as one
    consistent moment and is not one. Blank is the honest answer."""
    old = tmp_path / "perf_df_2020-01-01.pkl"
    old.write_bytes(b"not-a-real-pickle")
    os.utime(old, (0, 0))  # epoch: unambiguously stale
    monkeypatch.setattr(b, "REPO", str(tmp_path))
    monkeypatch.setattr(b, "SNAPSHOT_MAX_AGE_DAYS", 10)
    cache = tmp_path / "cache" / "perf"
    cache.mkdir(parents=True)
    stale = cache / "perf_df_2020-01-01.pkl"
    stale.write_bytes(b"not-a-real-pickle")
    os.utime(stale, (0, 0))
    returns, as_of = b.load_returns()
    assert returns == {} and as_of is None, (
        "a stale snapshot must produce blank returns, never stale ones read as live")


def test_no_snapshot_at_all_is_blank_not_a_crash(monkeypatch, tmp_path):
    monkeypatch.setattr(b, "REPO", str(tmp_path))
    assert b.load_returns() == ({}, None)


def test_a_workbook_open_in_excel_exits_3_not_1(monkeypatch, tmp_path):
    """Exit 3 means "JP has the file open", and the weekly build treats it as a
    warning rather than a red task — a red that fires because someone was reading
    the output trains you to ignore reds.

    The lock is simulated at `shutil.move`, NOT with a read-only attribute. The
    first attempt to verify this used `chmod 444` and the build exited 0, because
    `archive_existing` MOVES the old file out of the way and a move succeeds on a
    read-only file. Excel's is a sharing lock, so the failure lands on the move —
    which is exactly why the handler had to cover the archive step and not just
    the copy. A test that reproduces the wrong failure proves nothing.
    """
    current = tmp_path / ("%s.xlsx" % b.STEM)
    current.write_bytes(b"pretend workbook")

    def locked(*a, **k):
        raise PermissionError(32, "The process cannot access the file")

    row = {c: None for c in b.COLS}
    row.update({"Ticker": "ACME", "Company Name": "Acme Inc", "Sector": "MedTech",
                "Subsector": "Dental", "Mkt Cap (USD $M)": 1000.0,
                "Price (local)": 10.0, "Ccy": "USD"})
    monkeypatch.setattr(b, "build_records",
                        lambda asof: ([row], datetime.date(2026, 8, 21), []))
    monkeypatch.setattr(b.shutil, "move", locked)
    monkeypatch.setattr(b.shutil, "copy2", locked)
    monkeypatch.setattr(b, "PUBLIC_CSV", str(tmp_path / "docs" / "hc_coverage.csv"))
    monkeypatch.setattr("sys.argv", ["build", "--out-dir", str(tmp_path)])

    with pytest.raises(SystemExit) as e:
        b.main()
    assert e.value.code == 3, (
        "a workbook locked by Excel exited %r; the weekly build reads anything "
        "other than 3 as a pipeline failure and turns the task red"
        % (e.value.code,))


def test_the_stem_and_the_published_endpoint_are_decoupled():
    """The Sheet is one =IMPORTDATA() cell pointed at this exact URL and nothing
    in this repo can rewrite that cell, so renaming the workbook must never rename
    the published CSV."""
    assert b.STEM == "AA_Core Coverage"
    assert os.path.basename(b.PUBLIC_CSV) == "hc_coverage.csv"


def test_the_ratings_workbook_is_scoped_to_core_coverage():
    assert os.path.basename(b.RATINGS_PATH) == "Ratings_CoreCoverage.xlsx"
    assert b.HUMAN_COLS == {"Rating", "Notes"}
    assert not (b.HUMAN_COLS & b.MACHINE_COLS), \
        "a column cannot be both human-owned and machine-refreshed"


# ── precision and colour ─────────────────────────────────────────────────────

def test_every_column_with_a_number_format_also_declares_its_decimals():
    """The workbook's number format and the CSV's rounding read from two tables.
    If they drift, the same column shows 22 in one surface and 21.94 in the other
    and there is nothing to say which is intended."""
    assert set(b.NUMFMT) == set(b.DECIMALS), (
        set(b.NUMFMT) ^ set(b.DECIMALS))
    for col, fmt in b.NUMFMT.items():
        want_decimals = b.DECIMALS[col]
        has_decimals = "." in fmt
        assert has_decimals == (want_decimals > 0), (
            "%s: number format %r disagrees with DECIMALS=%d" % (col, fmt, want_decimals))


@pytest.mark.parametrize("col", ["Fwd P/E", "2019", "2024", "YTD",
                                 "Mkt Cap (USD $M)"])
def test_the_columns_jp_asked_to_lose_decimals_have_none(col):
    """JP 2026-08-26: "the annual returns dont need decimal point precision. and
    the Fwd P/e doesnt as well"."""
    assert b.DECIMALS[col] == 0
    assert "." not in b.NUMFMT[col]


def test_market_cap_shows_thousands_separators():
    assert "," in b.NUMFMT["Mkt Cap (USD $M)"]


def test_price_keeps_its_cents():
    """Rounding a price to whole units would make every sub-dollar name read 0."""
    assert b.DECIMALS["Price (local)"] == 2


def test_every_return_column_gets_its_own_colour_scale():
    """Per column, not one scale across the block: 2022 and 2021 have wildly
    different ranges and a shared gradient would render one of them flat."""
    # Newest of OUR outputs, whatever today's date stamp is -- pinning the plain
    # stem made both of these skip silently the moment filenames gained a date.
    import glob as _g
    hits = sorted(_g.glob(os.path.join(b.DEFAULT_OUT, "%s*.xlsx" % b.STEM)),
                  key=os.path.getmtime)
    path = hits[-1] if hits else ""
    if not os.path.exists(path):
        pytest.skip("workbook not built in this checkout")
    ws = openpyxl.load_workbook(path)["Coverage List"]
    hdr = [c for c in next(ws.iter_rows(min_row=4, max_row=4, values_only=True))]
    ranges = {str(r) for r in ws.conditional_formatting._cf_rules}
    assert len(ranges) == len(b.RETURN_COLS), (
        "expected one colour scale per return column, got %d for %d columns"
        % (len(ranges), len(b.RETURN_COLS)))
    for rng in ws.conditional_formatting._cf_rules:
        rules = ws.conditional_formatting._cf_rules[rng]
        assert [r.type for r in rules] == ["colorScale"], rules


def test_the_colour_scale_puts_white_at_zero_not_at_the_median():
    """A percentile midpoint would paint a column where everything fell as though
    half of it were fine. Anchoring white at 0 means the colour always answers
    "did this make money" and only the intensity is relative."""
    # Newest of OUR outputs, whatever today's date stamp is -- pinning the plain
    # stem made both of these skip silently the moment filenames gained a date.
    import glob as _g
    hits = sorted(_g.glob(os.path.join(b.DEFAULT_OUT, "%s*.xlsx" % b.STEM)),
                  key=os.path.getmtime)
    path = hits[-1] if hits else ""
    if not os.path.exists(path):
        pytest.skip("workbook not built in this checkout")
    ws = openpyxl.load_workbook(path)["Coverage List"]
    rng = next(iter(ws.conditional_formatting._cf_rules))
    rule = ws.conditional_formatting._cf_rules[rng][0]
    kinds = [c.type for c in rule.colorScale.cfvo]
    assert kinds == ["min", "num", "max"], kinds
    assert str(rule.colorScale.cfvo[1].val) in ("0", "0.0"), rule.colorScale.cfvo[1].val


# ── archiving must never touch JP's own files ────────────────────────────────

@pytest.mark.parametrize("name", [
    "Jason Peterson Coverage.xlsx",
    "Jason Peterson Coverage - ENIX.xlsx",
    "Coverage - LC Svcs, Medtech, JP coverage.xlsx",
    "Coverage - HC Services and MedTech - 2026-08-25 - ENIX.xlsx",
    "AA_Core Coverage NOTES.xlsx",     # his note, deliberately near-miss
    "AA_Core Coverage auto-updated.xlsx",   # no date -> not one of ours
    "my scratch.csv",
])
def test_archiving_leaves_jps_files_alone(tmp_path, name):
    """JP 2026-08-26: "I might put my own files in this coverage folder ... Don't
    move my files. You just archive the files you auto-generate."

    The lazy implementation archives every xlsx in the folder, which would file
    his work under archive/ and leave him unable to tell that from something he
    had misplaced himself. Matching is an explicit allow-list of names this script
    produces, and the near-miss cases above are the ones that would break a
    sloppier pattern."""
    (tmp_path / name).write_bytes(b"jp's file")
    moved = b.archive_previous_autogenerated(str(tmp_path), keep_names=[])
    assert moved == [], "archived a file that is not ours: %r" % moved
    assert (tmp_path / name).exists(), "%s was moved out from under JP" % name


@pytest.mark.parametrize("name", [
    "AA_Core Coverage auto-updated - 08.19.26.xlsx",
    "AA_Core Coverage auto-updated - 08.19.26.csv",
    "AA_Core Coverage.xlsx",
    "Coverage - HC Services and MedTech.xlsx",
])
def test_archiving_does_collect_our_own_earlier_output(tmp_path, name):
    (tmp_path / name).write_bytes(b"ours")
    moved = b.archive_previous_autogenerated(str(tmp_path), keep_names=[])
    assert moved == [name]
    assert not (tmp_path / name).exists()
    assert (tmp_path / "archive" / name).exists()


def test_todays_output_is_not_archived_by_its_own_run(tmp_path):
    """The date is in the filename now, so a same-day rebuild would otherwise file
    away the very workbook it is about to write."""
    today = "%s.xlsx" % b.dated_stem()
    (tmp_path / today).write_bytes(b"today")
    moved = b.archive_previous_autogenerated(str(tmp_path), keep_names=[today])
    assert moved == []
    assert (tmp_path / today).exists()


def test_dated_stem_uses_jps_format():
    """JP asked for "auto-updated - 08.21.26"."""
    assert b.dated_stem(datetime.date(2026, 8, 21)) == \
        "AA_Core Coverage auto-updated - 08.21.26"


# ── annualised returns ───────────────────────────────────────────────────────

def test_annualise_converts_a_cumulative_return():
    """The snapshot's 3Y/5Y are CUMULATIVE — calc_period_return(hist, 365*3).
    JNJ's 74.4 means 74% across three years, not per year; reporting it raw under
    a heading saying "annual" overstates it roughly threefold."""
    assert round(b.annualise(74.4, 3), 1) == 20.4
    assert round(b.annualise(71.5, 5), 1) == 11.4


def test_annualise_is_identity_over_one_year():
    assert round(b.annualise(12.0, 1), 6) == 12.0


def test_annualise_handles_losses():
    """-50% over 3 years is about -20% a year, not -16.7%."""
    assert round(b.annualise(-50.0, 3), 1) == -20.6


@pytest.mark.parametrize("cum", [-100.0, -150.0, None, ""])
def test_annualise_refuses_the_impossible(cum):
    """A security cannot lose more than everything, and the cube root of a
    negative growth factor is not a real number. Blank beats a made-up figure in a
    column people rank on."""
    assert b.annualise(cum, 3) is None


# ── the new layout ───────────────────────────────────────────────────────────

def test_column_order_is_jps():
    assert b.COLS[:6] == ["Ticker", "Company Name", "Rating",
                          "Mkt Cap (USD $M)", "EV (USD $M)", "Size"]
    assert b.COLS.index("Size") < b.COLS.index("Sector")
    assert b.COLS[-3:] == ["Listing", "Exchange", "Country (HQ)"]


def test_returns_run_most_recent_first():
    """JP: "the left most performance column should be YTD, and then 2025 and then
    the last performance column should be 2019"."""
    assert b.CALENDAR_RETURNS == ["YTD", "2025", "2024", "2023",
                                  "2022", "2021", "2020", "2019"]
    years = [c for c in b.CALENDAR_RETURNS if c != "YTD"]
    assert years == sorted(years, reverse=True)


def test_the_annualised_columns_say_they_are_annualised():
    """A heading of plain "3Y" beside calendar years would read as cumulative,
    which is exactly what the underlying snapshot column is."""
    for c in b.ANNUALISED_RETURNS:
        assert "ann" in c.lower()
    assert set(b.RETURN_COLS) == set(b.CALENDAR_RETURNS) | set(b.ANNUALISED_RETURNS)
